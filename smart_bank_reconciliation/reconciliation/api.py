import json
import frappe
from frappe import _
from frappe.utils import nowdate, getdate, flt, add_days
from .matching_engine import BankMatchingEngine

# The Select field also allows "Pending" (its default/unscored state) and blank/NULL for
# transactions never run through the matching engine. Any value outside this known-outcome
# set must fold into "Unmatched" so tile counts always sum to the total record count.
_KNOWN_RECON_QUEUES = {"Auto", "Review", "Unmatched", "High-Val", "Duplicate", "Aging", "Reconciled"}


def _suggested_entry_from_draft(raw):
    """The voucher the AI matched but deliberately did NOT store as this
    transaction's reconcilable entry — returned for display only.

    An invoice match stores no matched entry (a bank line cannot be cleared
    against an invoice directly; see approve_match), which left the "Matched
    ERP Entry" column blank even though the Actions modal happily showed the
    invoice — so the column and the modal appeared to disagree. The draft
    payload already records exactly which invoice it was, so read it back from
    there rather than storing an entry that must never be reconciled against.
    """
    if not raw:
        return None, None
    import json as _j
    try:
        payload = _j.loads(raw) if isinstance(raw, str) else raw
    except Exception:
        return None, None
    refs = (payload or {}).get("references") or []
    if not refs:
        return None, None
    first = refs[0] or {}
    return first.get("reference_name"), first.get("reference_doctype")


def _tally_queue_counts(rows, queue_field="recon_queue"):
    """Tile counts, one per raw bank statement line.

    These deliberately do NOT collapse consolidated groups. The table below
    renders a group as a single row, so the tiles and the row count differ on
    a statement with consolidations — that is intended: the tiles answer "how
    many bank transactions are in this period", which is the number that has
    to agree with the Bank Transaction list, while the table answers "how many
    things do I have to action".
    """
    from collections import Counter
    counts = Counter(
        r.get(queue_field) if r.get(queue_field) in _KNOWN_RECON_QUEUES else "Unmatched"
        for r in rows
    )
    return {
        "total":      len(rows),
        "auto":       counts.get("Auto", 0),
        "review":     counts.get("Review", 0),
        "unmatched":  counts.get("Unmatched", 0),
        "high_val":   counts.get("High-Val", 0),
        "duplicate":  counts.get("Duplicate", 0),
        "aging":      counts.get("Aging", 0),
        "reconciled": counts.get("Reconciled", 0),
    }


@frappe.whitelist()
def get_bank_transactions(bank_account, from_date, to_date):
    """Fetch raw bank transactions for the initial table view — no AI scoring."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    try:
        rows = frappe.db.get_all(
            "Bank Transaction",
            filters={"bank_account": bank_account, "date": ["between", [from_date, to_date]], "docstatus": 1},
            fields=["name", "date", "deposit", "withdrawal", "description",
                    "reference_number", "party_type", "party", "status", "unallocated_amount",
                    "recon_queue", "recon_confidence", "recon_matched_entries", "recon_match_type",
                    "recon_run_id", "recon_draft_payload"],
            order_by="date desc",
        )
    except Exception:
        rows = frappe.db.get_all(
            "Bank Transaction",
            filters={"bank_account": bank_account, "date": ["between", [from_date, to_date]], "docstatus": 1},
            fields=["name", "date", "deposit", "withdrawal", "description",
                    "reference_number", "party_type", "party", "status", "unallocated_amount"],
            order_by="date desc",
        )

    # A fully reconciled transaction belongs in the Reconciled queue and
    # nowhere else. recon_queue can disagree: it holds whatever the last AI
    # run decided (e.g. "Review"), and reconciling through ERPNext's own tools
    # — or any path that doesn't re-run matching — updates status/allocation
    # without touching it. The matching engine already applies this exact
    # precedence (status/zero-unallocated wins over any score) at the top of
    # its run loop; apply it here too so the plain, pre-AI table load and its
    # tile counts don't show reconciled rows under Review/Auto/Unmatched.
    for row in rows:
        if row.get("status") == "Reconciled" or float(row.get("unallocated_amount") or 0) == 0:
            row["recon_queue"] = "Reconciled"

    # queue_counts.total already dedupes consolidated groups (they display as
    # one row); keep the plain "total" field in lockstep so nothing that
    # reads it separately drifts out of sync with the tile counts.
    queue_counts = _tally_queue_counts(rows)
    total = queue_counts["total"]

    # recon_matched_entries is a plain stored list of names with no link
    # integrity behind it — nothing clears it when the referenced voucher is
    # later deleted or cancelled in ERP. Those stale names kept rendering in
    # the "Matched ERP Entry" column as links to vouchers that no longer
    # exist. Resolve every referenced name once, then drop the ones that
    # aren't a live submitted document.
    import json as _json

    def _referenced_entry_names(row):
        raw = row.get("recon_matched_entries")
        if not raw:
            return []
        try:
            return [n for n in (_json.loads(raw) or []) if n]
        except Exception:
            return []

    all_entry_names = set()
    for row in rows:
        all_entry_names.update(_referenced_entry_names(row))

    live_entries = {}   # name -> party_type ("" when the doctype has no party)
    if all_entry_names:
        name_list = list(all_entry_names)
        for pe in frappe.db.get_all(
            "Payment Entry",
            filters={"name": ["in", name_list], "docstatus": 1},
            fields=["name", "party_type"],
        ):
            live_entries[pe["name"]] = pe.get("party_type") or ""
            
        for doctype in ["Journal Entry", "Sales Invoice", "Purchase Invoice"]:
            for doc in frappe.db.get_all(
                doctype,
                filters={"name": ["in", name_list], "docstatus": 1},
                fields=["name"],
            ):
                live_entries.setdefault(doc["name"], "")
                
        remaining = [n for n in name_list if n not in live_entries]
        if remaining:
            for gl in frappe.db.get_all(
                "GL Entry",
                filters={"voucher_no": ["in", remaining], "is_cancelled": 0},
                fields=["voucher_no"],
                distinct=True,
            ):
                live_entries.setdefault(gl["voucher_no"], "")

        for row in rows:
            referenced = _referenced_entry_names(row)
            if not referenced:
                continue
            surviving = [n for n in referenced if n in live_entries]
            if len(surviving) != len(referenced):
                # At least one referenced voucher is gone/cancelled — rewrite the
                # field so the UI only ever shows what actually still exists.
                row["recon_matched_entries"] = frappe.as_json(surviving) if surviving else ""
            # Enrich party_type from the matched Payment Entry for transactions
            # with no party of their own (typical for imported statements).
            if not row.get("party_type"):
                for n in surviving:
                    if live_entries.get(n):
                        row["party_type"] = live_entries[n]
                        break

    # Display-only suggested entry for rows that store no reconcilable match,
    # derived here so a plain reload shows the same thing an AI run does. The
    # raw draft payload is then dropped — it is only needed to derive this, and
    # one JSON blob per row would bloat a 700-row statement for nothing.
    for row in rows:
        if not row.get("recon_matched_entries"):
            name, doctype = _suggested_entry_from_draft(row.get("recon_draft_payload"))
            if name:
                row["recon_suggested_entry"] = name
                row["recon_suggested_doctype"] = doctype or ""
        row.pop("recon_draft_payload", None)

    return {"transactions": rows, "total": total, "queue_counts": queue_counts}


@frappe.whitelist()
def resolve_entry_doctypes(names):
    """Bulk-resolve which doctype each matched ERP entry name belongs to, for
    building correct navigation links. recon_matched_entries only stores bare
    names, and a company's custom naming series (e.g. "SUM-INV-1160") can't be
    reliably guessed from the string alone — so this looks the names up directly
    instead of pattern-matching on them."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    if isinstance(names, str):
        names = json.loads(names)
    remaining = {n for n in (names or []) if n}
    if not remaining:
        return {}

    result = {}
    for doctype, route in (
        ("Journal Entry", "journal-entry"),
        ("Payment Entry", "payment-entry"),
        ("Sales Invoice", "sales-invoice"),
        ("Purchase Invoice", "purchase-invoice"),
    ):
        if not remaining:
            break
        found = frappe.db.get_all(doctype, filters={"name": ["in", list(remaining)]}, pluck="name")
        for n in found:
            result[n] = route
            remaining.discard(n)
    return result


@frappe.whitelist()
def get_suggestions(bank_account, from_date, to_date, company, settings_json=None):
    """Run the matching engine and return suggestions for the UI."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    settings = None
    if settings_json:
        try:
            settings = json.loads(settings_json) if isinstance(settings_json, str) else settings_json
        except Exception:
            pass
    if not settings:
        settings = _load_sbr_settings()

    engine = BankMatchingEngine(bank_account, from_date, to_date, company, settings=settings)
    results = engine.run()
    queue_counts = engine.get_queue_counts(results)

    # Build suggestion payload for each non-reconciled transaction
    suggestions = []
    for txn in results:
        if txn.get("recon_queue") in ("Reconciled", "Unmatched") or not txn.get("recon_queue"):
            continue
        suggestions.append({
            "bank_txn": txn["name"],
            "date": str(txn.get("date") or ""),
            "deposit": txn.get("deposit") or 0,
            "withdrawal": txn.get("withdrawal") or 0,
            "description": txn.get("description") or "",
            "reference_number": txn.get("reference_number") or "",
            "party": txn.get("party") or "",
            "queue": txn.get("recon_queue"),
            "confidence": txn.get("recon_confidence") or 0,
            "match_type": txn.get("recon_match_type") or "",
            "reasoning": txn.get("recon_ai_reasoning") or "",
            "matched": txn.get("matched"),
            "draft_payload": txn.get("recon_draft_payload"),
            "duplicate_of": txn.get("recon_duplicate_of") or [],
        })

    return {
        "queue_counts": queue_counts,
        "transactions": [
            {
                "name": t["name"],
                "date": str(t.get("date") or ""),
                "deposit": t.get("deposit") or 0,
                "withdrawal": t.get("withdrawal") or 0,
                "description": t.get("description") or "",
                "reference_number": t.get("reference_number") or "",
                "party_type": t.get("party_type") or (t.get("matched") or {}).get("party_type") or "",
                "party": t.get("party") or (t.get("matched") or {}).get("party") or "",
                "status": t.get("status") or "",
                "unallocated_amount": t.get("unallocated_amount") or 0,
                "recon_queue": t.get("recon_queue"),
                "recon_confidence": t.get("recon_confidence") or 0,
                "recon_matched_entries": t.get("recon_matched_entries") or "",
                "recon_match_type": t.get("recon_match_type") or "",
                # Consolidate group key — without it the table can't collapse a
                # consolidated group back into its single combined row (and its
                # combined total) after an AI Match run, the way a plain reload
                # via get_bank_transactions does.
                "recon_run_id": t.get("recon_run_id") or "",
                # Display-only: the voucher the AI matched but that is not
                # stored as a reconcilable entry (invoice matches). Without
                # this the column read "—" while Actions showed the invoice.
                "recon_suggested_entry": _suggested_entry_from_draft(
                    t.get("recon_draft_payload"))[0] or "",
                "recon_suggested_doctype": _suggested_entry_from_draft(
                    t.get("recon_draft_payload"))[1] or "",
                "recon_duplicate_of": t.get("recon_duplicate_of") or [],
            }
            for t in results
        ],
        "suggestions": suggestions,
    }


@frappe.whitelist()
def start_recon_job(bank_account, from_date, to_date, company, settings_json=None):
    """Enqueue a background AI matching run and return the job key for polling."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    import hashlib

    job_key = "sbr_job_" + hashlib.md5(
        f"{bank_account}|{from_date}|{to_date}".encode()
    ).hexdigest()[:12]

    # Mark as running before enqueuing so the first poll sees the right state
    frappe.cache().set_value(job_key, {"status": "running"}, expires_in_sec=3600)

    frappe.enqueue(
        "smart_bank_reconciliation.reconciliation.api._run_recon_job_bg",
        queue="default",
        timeout=600,
        job_key=job_key,
        bank_account=bank_account,
        from_date=from_date,
        to_date=to_date,
        company=company,
        settings_json=settings_json,
    )

    return {"job_key": job_key}


def _run_recon_job_bg(job_key, bank_account, from_date, to_date, company, settings_json=None):
    """Background worker called by start_recon_job. Stores result in Redis."""
    try:
        settings = None
        if settings_json:
            try:
                settings = json.loads(settings_json) if isinstance(settings_json, str) else settings_json
            except Exception:
                pass
        if not settings:
            settings = _load_sbr_settings()

        engine = BankMatchingEngine(bank_account, from_date, to_date, company, settings=settings)
        results = engine.run()
        queue_counts = engine.get_queue_counts(results)

        suggestions = []
        for txn in results:
            if txn.get("recon_queue") in ("Reconciled", "Unmatched") or not txn.get("recon_queue"):
                continue
            suggestions.append({
                "bank_txn": txn["name"],
                "date": str(txn.get("date") or ""),
                "deposit": txn.get("deposit") or 0,
                "withdrawal": txn.get("withdrawal") or 0,
                "description": txn.get("description") or "",
                "reference_number": txn.get("reference_number") or "",
                "party": txn.get("party") or "",
                "queue": txn.get("recon_queue"),
                "confidence": txn.get("recon_confidence") or 0,
                "match_type": txn.get("recon_match_type") or "",
                "reasoning": txn.get("recon_ai_reasoning") or "",
                "matched": txn.get("matched"),
                "draft_payload": txn.get("recon_draft_payload"),
                "duplicate_of": txn.get("recon_duplicate_of") or [],
            })

        data = {
            "queue_counts": queue_counts,
            "transactions": [
                {
                    "name": t["name"],
                    "date": str(t.get("date") or ""),
                    "deposit": t.get("deposit") or 0,
                    "withdrawal": t.get("withdrawal") or 0,
                    "description": t.get("description") or "",
                    "reference_number": t.get("reference_number") or "",
                    "party_type": t.get("party_type") or (t.get("matched") or {}).get("party_type") or "",
                    "party": t.get("party") or "",
                    "status": t.get("status") or "",
                    "unallocated_amount": t.get("unallocated_amount") or 0,
                    "recon_queue": t.get("recon_queue"),
                    "recon_confidence": t.get("recon_confidence") or 0,
                    "recon_matched_entries": t.get("recon_matched_entries") or "",
                    "recon_match_type": t.get("recon_match_type") or "",
                    # Consolidate group key — same reason as the synchronous
                    # payload above. This is the background-job path, so
                    # omitting it broke consolidated groups apart specifically
                    # when the match ran async.
                    "recon_run_id": t.get("recon_run_id") or "",
                    # Display-only: the voucher the AI matched but that is not
                    # stored as a reconcilable entry (invoice matches). Without
                    # this the column read "—" while Actions showed the invoice.
                    "recon_suggested_entry": _suggested_entry_from_draft(
                        t.get("recon_draft_payload"))[0] or "",
                    "recon_suggested_doctype": _suggested_entry_from_draft(
                        t.get("recon_draft_payload"))[1] or "",
                    "recon_duplicate_of": t.get("recon_duplicate_of") or [],
                }
                for t in results
            ],
            "suggestions": suggestions,
        }

        frappe.cache().set_value(
            job_key,
            {"status": "complete", "data": data},
            expires_in_sec=3600,
        )
    except Exception:
        traceback_str = frappe.get_traceback()
        frappe.log_error(title="AI Matching Job Failed", message=traceback_str)
        frappe.cache().set_value(
            job_key,
            {"status": "error", "message": traceback_str},
            expires_in_sec=600,
        )


@frappe.whitelist()
def get_recon_job_status(job_key):
    """Poll the status of a background reconciliation job."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    result = frappe.cache().get_value(job_key)
    if not result:
        return {"status": "expired"}
    # Don't return full data payload in running/error states
    if result.get("status") != "complete":
        return {"status": result.get("status"), "message": result.get("message")}
    return result


@frappe.whitelist()
def approve_match(bank_transaction, matched_entries, match_type=None):
    """
    Set clearance_date on matched PE/JE entries and mark the Bank Transaction
    as Reconciled.
    """
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    if isinstance(matched_entries, str):
        import json
        matched_entries = json.loads(matched_entries)

    clearance_date = nowdate()

    txn_data = frappe.db.get_value(
        "Bank Transaction", bank_transaction,
        ["deposit", "withdrawal", "unallocated_amount", "party", "description"],
        as_dict=True,
    )
    bank_amount = float(txn_data.get("deposit") or txn_data.get("withdrawal") or 0)

    # Support both old format (list of name strings) and new format (list of {name, amount} dicts)
    def _entry_name(e):
        return e["name"] if isinstance(e, dict) else e

    def _alloc_amount(e, n_entries):
        if isinstance(e, dict) and e.get("amount"):
            return float(e["amount"])
        return bank_amount / n_entries

    import time as _time

    def _update_with_retry(table, name, date, max_retries=3):
        # Retry on MySQL 1020 (record changed) / 1213 (deadlock) to survive concurrent writes.
        for attempt in range(max_retries):
            try:
                frappe.db.sql(
                    "UPDATE `{0}` SET `clearance_date`=%s WHERE `name`=%s".format(table),
                    (date, name),
                )
                return
            except Exception as exc:
                err = str(exc)
                if attempt < max_retries - 1 and any(c in err for c in ("1020", "1213", "Deadlock", "Lock wait")):
                    _time.sleep(0.15 * (attempt + 1))
                    frappe.db.rollback()
                    continue
                raise

    n = len(matched_entries)
    # Only Payment Entry and Journal Entry can actually be cleared and linked.
    # Anything else the user managed to pick (an Expense Claim or other voucher
    # surfaced by the GL sweep in get_erp_vouchers_for_match) resolves to
    # nothing — track those so the tail of this function can refuse rather than
    # marking the transaction Reconciled with no reference row behind it.
    linked_any = False
    unresolved = []
    for entry in matched_entries:
        entry_name = _entry_name(entry)
        alloc = _alloc_amount(entry, n)

        if frappe.db.exists("Payment Entry", entry_name):
            table, payment_document = "tabPayment Entry", "Payment Entry"
        elif frappe.db.exists("Journal Entry", entry_name):
            table, payment_document = "tabJournal Entry", "Journal Entry"
        elif frappe.db.exists("Sales Invoice", entry_name) or frappe.db.exists("Purchase Invoice", entry_name):
            frappe.throw(
                "Cannot reconcile bank transaction directly against invoice {0}. "
                "Create a Payment Entry for the invoice first, then reconcile against the Payment Entry.".format(entry_name)
            )
        else:
            unresolved.append(entry_name)
            continue

        already_linked_here = frappe.db.exists(
            "Bank Transaction Payments",
            {"parent": bank_transaction, "payment_entry": entry_name},
        )

        # Guard against double-allocating one ERP entry to two different bank
        # transactions — e.g. two transactions independently top-scoring the
        # same Payment Entry within a single bulk-approve batch. If this
        # entry is already cleared against a DIFFERENT bank transaction,
        # refuse rather than silently linking a second one to the same money
        # — UNLESS that other transaction is a sibling in the same Consolidate
        # group as this one (same recon_match_type="Consolidated" AND the
        # same recon_run_id, repurposed as the group key — see
        # _consolidate_via_existing_match), which deliberately shares one
        # matched entry across every member on purpose.
        # payment_document ("Payment Entry"), NOT table ("tabPayment Entry"):
        # get_value takes a DocType and prepends "tab" itself, so passing the
        # raw table name queried "tabtabPayment Entry", errored, and returned
        # None — silently disabling this whole double-allocation guard.
        existing_clearance = frappe.db.get_value(payment_document, entry_name, "clearance_date")
        if existing_clearance and not already_linked_here:
            linked_txns = frappe.db.get_all(
                "Bank Transaction Payments",
                filters={"payment_entry": entry_name, "parenttype": "Bank Transaction"},
                pluck="parent",
            )
            my_txn = frappe.db.get_value(
                "Bank Transaction", bank_transaction,
                ["recon_match_type", "recon_run_id"], as_dict=True,
            )
            is_same_group = False
            if my_txn and my_txn.recon_match_type == "Consolidated" and my_txn.recon_run_id:
                for lt in linked_txns:
                    other = frappe.db.get_value(
                        "Bank Transaction", lt,
                        ["recon_match_type", "recon_run_id"], as_dict=True,
                    )
                    if (other and other.recon_match_type == "Consolidated"
                            and other.recon_run_id == my_txn.recon_run_id):
                        is_same_group = True
                        break
            if not is_same_group:
                frappe.throw(
                    "{0} {1} is already reconciled (cleared {2}) against another bank "
                    "transaction. Un-reconcile it there first before matching it here.".format(
                        payment_document, entry_name, existing_clearance
                    )
                )

        _update_with_retry(table, entry_name, clearance_date)
        linked_any = True

        # Insert into Bank Transaction Payments child table (ERPNext native linking)
        if not already_linked_here:
            frappe.db.sql(
                """INSERT INTO `tabBank Transaction Payments`
                   (name, creation, modified, modified_by, owner, docstatus,
                    parent, parentfield, parenttype, idx,
                    payment_document, payment_entry, allocated_amount, clearance_date)
                   VALUES (%s, NOW(), NOW(), %s, %s, 0,
                           %s, 'payment_entries', 'Bank Transaction', 1,
                           %s, %s, %s, %s)""",
                (
                    frappe.generate_hash(length=10),
                    frappe.session.user, frappe.session.user,
                    bank_transaction,
                    payment_document, entry_name,
                    alloc,
                    clearance_date,
                ),
            )

    # If the caller supplied entries but not one of them could be cleared and
    # linked, refuse. Previously this fell through and still marked the
    # transaction fully Reconciled with unallocated_amount 0 — leaving a
    # reconciled bank line with no reference row behind it and no error shown.
    #
    # An EMPTY matched_entries list is deliberately still allowed: that is the
    # modal's "Mark as Reconciled without linking an ERP voucher" action, which
    # is a supported way to close out a line (bank charges, opening balances).
    # Only a caller that asked for specific entries and got none of them linked
    # is an error.
    if matched_entries and not linked_any:
        frappe.throw(
            "Could not reconcile {0}: {1} cannot be linked to a bank transaction. "
            "Only Payment Entries and Journal Entries can be cleared directly — "
            "create a Payment Entry for this voucher and reconcile against that "
            "instead.".format(bank_transaction, ", ".join(unresolved) or "the selected voucher")
        )

    # Mark bank transaction as reconciled
    frappe.db.set_value("Bank Transaction", bank_transaction, {
        "recon_queue": "Reconciled",
        "recon_user_action": "Accepted",
        "status": "Reconciled",
        "unallocated_amount": 0,
    })

    # Record approval pattern so future AI runs benefit from this feedback
    txn_fields = frappe.db.get_value(
        "Bank Transaction", bank_transaction, ["party", "description"], as_dict=True
    )
    if txn_fields:
        from .pattern_store import PatternStore
        PatternStore().record_approval(
            txn_fields.get("party") or "",
            txn_fields.get("description") or "",
        )

    frappe.db.commit()

    return {"status": "ok", "clearance_date": clearance_date}


@frappe.whitelist()
def bulk_approve_auto(bank_account, from_date, to_date):
    """Approve all transactions currently in the Auto queue."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import json

    auto_txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "recon_queue": "Auto",
            "recon_user_action": ["in", ["Pending", ""]],
        },
        fields=["name", "recon_matched_entries"],
    )

    approved = []
    for txn in auto_txns:
        try:
            entries = json.loads(txn.get("recon_matched_entries") or "[]")
            approve_match(txn["name"], entries)
            approved.append(txn["name"])
        except Exception:
            frappe.log_error(frappe.get_traceback(),
                             f"Bulk approve failed for {txn['name']}")

    # Return updated counts
    all_txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
        },
        fields=["recon_queue", "recon_match_type", "recon_matched_entries", "recon_run_id"],
    )

    return {
        "count": len(approved),
        "approved_transactions": approved,
        "new_counts": _tally_queue_counts(all_txns),
    }


@frappe.whitelist()
def bulk_approve_transactions(transaction_names):
    """Approve a caller-supplied list of Bank Transactions in one call.

    Used by the Review/Aging tabs' "Bulk Approve Filtered" action, which
    approves exactly whatever the user currently has visible under their
    active filters (confidence range, aging range, search text, ...) —
    the frontend already knows that exact set, so this takes it directly
    rather than re-deriving a single filter dimension server-side.
    """
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import json
    if isinstance(transaction_names, str):
        transaction_names = json.loads(transaction_names)
    if not transaction_names:
        return {"count": 0, "approved_transactions": [], "skipped": []}

    txns = frappe.db.get_all(
        "Bank Transaction",
        filters={"name": ["in", transaction_names], "docstatus": 1},
        fields=["name", "recon_matched_entries", "recon_queue"],
    )

    approved, skipped = [], []
    for txn in txns:
        if txn.get("recon_queue") == "Reconciled":
            continue
        try:
            entries = json.loads(txn.get("recon_matched_entries") or "[]")
            if not entries:
                skipped.append(txn["name"])
                continue
            approve_match(txn["name"], entries)
            approved.append(txn["name"])
        except Exception:
            frappe.log_error(frappe.get_traceback(),
                             f"Bulk approve (filtered) failed for {txn['name']}")
            skipped.append(txn["name"])

    return {
        "count":                len(approved),
        "approved_transactions": approved,
        "skipped":              skipped,
    }


@frappe.whitelist()
def bulk_reject_review(bank_account, from_date, to_date):
    """Reject all transactions currently in the Review queue."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    review_txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "recon_queue": "Review",
            "recon_user_action": ["in", ["Pending", ""]],
        },
        fields=["name"],
    )

    rejected = []
    for txn in review_txns:
        frappe.db.set_value("Bank Transaction", txn["name"], {
            "recon_user_action": "Rejected",
            "recon_ai_reasoning": "Bulk rejected by user",
        })
        rejected.append(txn["name"])

    all_txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
        },
        fields=["recon_queue", "recon_match_type", "recon_matched_entries", "recon_run_id"],
    )

    return {
        "count": len(rejected),
        "rejected_transactions": rejected,
        "new_counts": _tally_queue_counts(all_txns),
    }


@frappe.whitelist()
def rerun_ai_on_transactions(transaction_names, bank_account, from_date, to_date, company):
    """Reset recon fields for selected transactions and re-run AI matching only on those."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import json
    if isinstance(transaction_names, str):
        transaction_names = json.loads(transaction_names)
    if not transaction_names:
        frappe.throw(_("No transactions selected."))

    # Bulk-reset recon fields (skip truly reconciled transactions)
    in_clause = ", ".join(["%s"] * len(transaction_names))
    frappe.db.sql(
        "UPDATE `tabBank Transaction` SET "
        "recon_queue = NULL, recon_confidence = 0, recon_matched_entries = NULL, "
        "recon_match_type = NULL, recon_run_id = NULL, recon_ai_reasoning = NULL, "
        "recon_draft_payload = NULL, recon_signals_json = NULL, recon_wht_amount = NULL, "
        "recon_user_action = NULL "
        "WHERE name IN (" + in_clause + ") AND status != 'Reconciled'",
        transaction_names,
    )

    settings = _load_sbr_settings()
    engine = BankMatchingEngine(
        bank_account, from_date, to_date, company,
        settings=settings, only_names=transaction_names,
    )
    results = engine.run()

    # Full period queue counts (reflects all transactions, not just selected)
    all_txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
        },
        fields=["recon_queue", "recon_match_type", "recon_matched_entries", "recon_run_id"],
    )
    queue_counts = _tally_queue_counts(all_txns)

    suggestions = []
    transactions_out = []
    for txn in results:
        if txn.get("recon_queue") in ("Reconciled", "Unmatched") or not txn.get("recon_queue"):
            continue
        suggestions.append({
            "bank_txn":        txn["name"],
            "date":            str(txn.get("date") or ""),
            "deposit":         txn.get("deposit") or 0,
            "withdrawal":      txn.get("withdrawal") or 0,
            "description":     txn.get("description") or "",
            "reference_number": txn.get("reference_number") or "",
            "party":           txn.get("party") or "",
            "queue":           txn.get("recon_queue"),
            "confidence":      txn.get("recon_confidence") or 0,
            "match_type":      txn.get("recon_match_type") or "",
            "reasoning":       txn.get("recon_ai_reasoning") or "",
            "matched":         txn.get("matched"),
            "draft_payload":   txn.get("recon_draft_payload"),
            "duplicate_of":    txn.get("recon_duplicate_of") or [],
        })
        transactions_out.append({
            "name":                  txn["name"],
            "date":                  str(txn.get("date") or ""),
            "deposit":               txn.get("deposit") or 0,
            "withdrawal":            txn.get("withdrawal") or 0,
            "description":           txn.get("description") or "",
            "reference_number":      txn.get("reference_number") or "",
            "party_type":            txn.get("party_type") or (txn.get("matched") or {}).get("party_type") or "",
            "party":                 txn.get("party") or "",
            "status":                txn.get("status") or "",
            "unallocated_amount":    txn.get("unallocated_amount") or 0,
            "recon_queue":           txn.get("recon_queue"),
            "recon_confidence":      txn.get("recon_confidence") or 0,
            "recon_matched_entries": txn.get("recon_matched_entries") or "",
            "recon_match_type":      txn.get("recon_match_type") or "",
            # Consolidate group key — carried on every transaction payload so a
            # consolidated group survives a targeted re-run the same as a full
            # match or a plain reload.
            "recon_run_id":          txn.get("recon_run_id") or "",
            # Display-only: the voucher the AI matched but that is not
            # stored as a reconcilable entry (invoice matches). Without
            # this the column read "—" while Actions showed the invoice.
            "recon_suggested_entry": _suggested_entry_from_draft(
                txn.get("recon_draft_payload"))[0] or "",
            "recon_suggested_doctype": _suggested_entry_from_draft(
                txn.get("recon_draft_payload"))[1] or "",
            "recon_duplicate_of":    txn.get("recon_duplicate_of") or [],
        })

    return {
        "rerun_count":  len(transaction_names),
        "queue_counts": queue_counts,
        "transactions": transactions_out,
        "suggestions":  suggestions,
    }


@frappe.whitelist()
def reject_suggestion(bank_transaction, rejection_reason=None):
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    frappe.db.commit()  # end read snapshot before write to avoid MySQL 1020
    frappe.db.set_value("Bank Transaction", bank_transaction, {
        "recon_user_action": "Rejected",
        "recon_ai_reasoning": rejection_reason or "",
    })

    # Record rejection pattern so future AI runs suppress this match
    txn_fields = frappe.db.get_value(
        "Bank Transaction", bank_transaction, ["party", "description"], as_dict=True
    )
    if txn_fields:
        from .pattern_store import PatternStore
        PatternStore().record_rejection(
            txn_fields.get("party") or "",
            txn_fields.get("description") or "",
        )

    return {"status": "ok"}


@frappe.whitelist()
def mark_duplicate_investigated(bank_transaction):
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    frappe.db.commit()  # end read snapshot before write to avoid MySQL 1020
    frappe.db.set_value("Bank Transaction", bank_transaction,
                        "recon_user_action", "Investigated")
    return {"status": "ok"}


@frappe.whitelist()
def delete_duplicate_transaction(bank_transaction):
    """Cancel and permanently delete a Bank Transaction in the Duplicate queue.
    After deletion the caller should re-run AI so the surviving twin gets a proper queue."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    recon_queue = frappe.db.get_value("Bank Transaction", bank_transaction, "recon_queue")
    if (recon_queue or "") != "Duplicate":
        frappe.throw(_("Only Duplicate-queue transactions can be deleted this way."))

    # Use direct SQL to cancel — avoids Frappe's check_if_latest() optimistic-lock (1020)
    frappe.db.sql(
        "UPDATE `tabBank Transaction` SET docstatus=2 WHERE name=%s AND docstatus=1",
        (bank_transaction,),
    )
    frappe.db.sql(
        "DELETE FROM `tabBank Transaction` WHERE name=%s",
        (bank_transaction,),
    )
    frappe.db.commit()

    return {"status": "deleted", "name": bank_transaction}


@frappe.whitelist()
def delete_reversed_transaction(bank_transaction):
    """Cancel and permanently delete a Bank Transaction the tool flagged as a
    reversal/bounce (recon_match_type "Reversal" — see nigerian_rules).

    Deliberately a separate endpoint from delete_duplicate_transaction rather
    than loosening that one's Duplicate-only guard: each stays narrow about what
    it will destroy. Reconciled lines are refused — unreconcile first, otherwise
    deleting would strip a bank line out from under a cleared ERP voucher.
    """
    frappe.only_for(["Accounts Manager", "System Manager"])

    row = frappe.db.get_value(
        "Bank Transaction", bank_transaction,
        ["recon_match_type", "status", "docstatus"], as_dict=True,
    )
    if not row:
        frappe.throw(_("Bank Transaction {0} not found.").format(bank_transaction))
    if (row.recon_match_type or "") != "Reversal":
        frappe.throw(_("Only transactions flagged as a Reversal can be deleted this way."))
    if (row.status or "") == "Reconciled":
        frappe.throw(_(
            "{0} is already reconciled. Un-reconcile it first, then delete."
        ).format(bank_transaction))

    # Direct SQL cancel + delete — avoids Frappe's check_if_latest() optimistic
    # lock (error 1020), same approach as the duplicate deletion above.
    frappe.db.sql(
        "UPDATE `tabBank Transaction` SET docstatus=2 WHERE name=%s AND docstatus=1",
        (bank_transaction,),
    )
    # Drop the allocation child rows first. A raw parent DELETE leaves them
    # orphaned, and ERPNext derives how much of a Payment Entry is still
    # allocatable from these rows — orphans would make a PE look permanently
    # consumed by a transaction that no longer exists, silently blocking it
    # from ever being matched again.
    frappe.db.sql(
        "DELETE FROM `tabBank Transaction Payments` WHERE parent=%s",
        (bank_transaction,),
    )
    frappe.db.sql("DELETE FROM `tabBank Transaction` WHERE name=%s", (bank_transaction,))
    frappe.db.commit()

    return {"status": "deleted", "name": bank_transaction}


@frappe.whitelist()
def bulk_delete_duplicate_transactions(bank_transactions):
    """Cancel and delete multiple Bank Transactions that are all in the Duplicate queue."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import json
    names = json.loads(bank_transactions) if isinstance(bank_transactions, str) else bank_transactions
    if not names:
        return {"deleted": [], "count": 0}

    # Fetch recon_queue for all in one query to avoid per-doc ORM loads
    rows = frappe.db.get_all(
        "Bank Transaction",
        filters={"name": ["in", names]},
        fields=["name", "recon_queue"],
    )
    queue_map = {r.name: (r.recon_queue or "") for r in rows}

    to_delete = [n for n in names if queue_map.get(n) == "Duplicate"]
    skipped   = [n for n in names if queue_map.get(n) != "Duplicate"]
    deleted   = []

    for name in to_delete:
        try:
            # Direct SQL cancel + delete — avoids check_if_latest() 1020 error
            frappe.db.sql(
                "UPDATE `tabBank Transaction` SET docstatus=2 WHERE name=%s AND docstatus=1",
                (name,),
            )
            frappe.db.sql(
                "DELETE FROM `tabBank Transaction` WHERE name=%s",
                (name,),
            )
            deleted.append(name)
        except Exception as e:
            skipped.append(name)
            frappe.log_error(f"bulk_delete_duplicate: {name} — {e}")

    frappe.db.commit()
    return {"deleted": deleted, "skipped": skipped, "count": len(deleted)}


@frappe.whitelist()
def escalate_to_ar(bank_transaction, note=None):
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    frappe.db.commit()  # end read snapshot before write to avoid MySQL 1020
    frappe.db.set_value("Bank Transaction", bank_transaction, {
        "recon_user_action": "Escalated",
        "recon_ai_reasoning": (note or "") + " [Escalated to AR/AP team]",
    })
    return {"status": "ok"}


@frappe.whitelist()
def approve_high_value(bank_transaction, approver_note=None):
    """First or second approval for High-Value transactions (>NGN 50M)."""
    frappe.only_for(["Accounts Manager", "System Manager"])

    txn = frappe.db.get_value(
        "Bank Transaction",
        bank_transaction,
        ["recon_approval_1", "recon_matched_entries", "recon_queue"],
        as_dict=True,
    )

    if txn.get("recon_queue") != "High-Val":
        frappe.throw(_("This transaction is not in the High-Value queue."))

    user = frappe.session.user

    if not txn.get("recon_approval_1"):
        frappe.db.commit()  # end read snapshot before write to avoid MySQL 1020
        frappe.db.set_value("Bank Transaction", bank_transaction, {
            "recon_approval_1": user,
        })
        return {"status": "first_approved", "message": f"First approval by {user} recorded."}
    else:
        if txn.get("recon_approval_1") == user:
            frappe.throw(_("Same user cannot provide both approvals."))
        import json
        entries = json.loads(txn.get("recon_matched_entries") or "[]")
        approve_match(bank_transaction, entries)
        return {"status": "fully_approved", "message": "High-value transaction reconciled."}


@frappe.whitelist()
def create_draft_entry(bank_transaction, entry_type, prefill):
    """Create a draft (unsaved) PE or JE from the pre-fill payload."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import json
    if isinstance(prefill, str):
        prefill = json.loads(prefill)

    # Resolve bank transaction data (needed by both PE and JE paths)
    bt_data = frappe.db.get_value(
        "Bank Transaction", bank_transaction,
        ["bank_account", "deposit", "withdrawal", "unallocated_amount"],
        as_dict=True,
    ) or {}
    bank_account_name = prefill.get("bank_account") or bt_data.get("bank_account") or ""
    ba_data = frappe.db.get_value(
        "Bank Account", bank_account_name, ["account", "company"], as_dict=True
    ) if bank_account_name else None
    gl_account   = (ba_data.account  if ba_data else None) or ""
    company_name = prefill.get("company") or (ba_data.company if ba_data else None) or \
                   frappe.defaults.get_defaults().get("company")

    if entry_type == "PE":
        dep = float(bt_data.get("deposit") or 0)
        paid_amount  = float(prefill.get("paid_amount") or prefill.get("received_amount") or
                             bt_data.get("unallocated_amount") or dep or
                             bt_data.get("withdrawal") or 0)
        payment_type = prefill.get("payment_type") or ("Receive" if dep > 0 else "Pay")

        doc = frappe.new_doc("Payment Entry")
        doc.payment_type    = payment_type
        doc.company         = company_name
        doc.posting_date    = prefill.get("posting_date", nowdate())
        doc.party_type      = prefill.get("party_type", "")
        doc.party           = prefill.get("party", "")
        doc.paid_amount     = paid_amount
        doc.received_amount = paid_amount
        doc.reference_no    = prefill.get("reference_no", "") or \
                              frappe.db.get_value("Bank Transaction", bank_transaction, "reference_number") or \
                              str(prefill.get("posting_date", nowdate()))
        doc.reference_date  = prefill.get("reference_date", "") or prefill.get("posting_date", "") or nowdate()
        doc.remarks         = prefill.get("remarks", "")
        if prefill.get("mode_of_payment"):
            doc.mode_of_payment = prefill["mode_of_payment"]
        if prefill.get("project"):
            doc.project = prefill["project"]
        if prefill.get("cost_center"):
            doc.cost_center = prefill["cost_center"]
        if payment_type == "Receive":
            doc.paid_to = gl_account
        else:
            doc.paid_from = gl_account
        for ref in prefill.get("references") or []:
            if ref.get("reference_doctype") and ref.get("reference_name"):
                inv_data = frappe.db.get_value(
                    ref["reference_doctype"], ref["reference_name"],
                    ["outstanding_amount", "grand_total", "due_date"],
                    as_dict=True,
                ) or {}
                doc.append("references", {
                    "reference_doctype":  ref["reference_doctype"],
                    "reference_name":     ref["reference_name"],
                    "allocated_amount":   float(ref.get("allocated_amount") or paid_amount),
                    "outstanding_amount": float(inv_data.get("outstanding_amount") or 0),
                    "total_amount":       float(inv_data.get("grand_total") or 0),
                    "due_date":           inv_data.get("due_date"),
                })
        doc.insert(ignore_permissions=True, ignore_mandatory=True)
    else:
        dep = float(bt_data.get("deposit") or 0)
        wit = float(bt_data.get("withdrawal") or 0)
        second_account = prefill.get("second_account", "")

        doc = frappe.new_doc("Journal Entry")
        doc.voucher_type  = prefill.get("voucher_type", "Bank Entry")
        doc.posting_date  = prefill.get("posting_date", nowdate())
        doc.cheque_no     = prefill.get("cheque_no", "") or prefill.get("reference_number", "")
        doc.cheque_date   = prefill.get("cheque_date", "") or prefill.get("reference_date", "")
        doc.remark        = prefill.get("narration", "") or prefill.get("remark", "")
        if prefill.get("mode_of_payment"):
            doc.mode_of_payment = prefill["mode_of_payment"]

        if second_account and gl_account:
            # Mirror V13's create_journal_entry_bts account wiring
            doc.append("accounts", {
                "account": second_account,
                "credit_in_account_currency": dep if dep > 0 else 0,
                "debit_in_account_currency":  wit if wit > 0 else 0,
                "party_type": prefill.get("party_type", ""),
                "party":      prefill.get("party", ""),
            })
            doc.append("accounts", {
                "account": gl_account,
                "bank_account": bank_account_name,
                "credit_in_account_currency": wit if wit > 0 else 0,
                "debit_in_account_currency":  dep if dep > 0 else 0,
            })
        else:
            amount = prefill.get("amount", 0)
            doc.append("accounts", {"account": prefill.get("debit_account", ""), "debit_in_account_currency": amount})
            doc.append("accounts", {"account": prefill.get("credit_account", ""), "credit_in_account_currency": amount})
        doc.insert(ignore_permissions=True, ignore_mandatory=True)

    return {"doctype": doc.doctype, "name": doc.name}


def _gl_balance(account, cutoff_date):
    """
    Return the exact GL balance for the account up to and including cutoff_date,
    using standard ERPNext get_balance_on so it perfectly matches the General Ledger report.
    """
    from erpnext.accounts.utils import get_balance_on
    balance = get_balance_on(account, getdate(cutoff_date))
    return float(balance or 0)


@frappe.whitelist()
def get_account_opening_balance(bank_account, from_date):
    """Return the GL balance for the day before from_date — matches the
    'Opening' row of the General Ledger report."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    try:
        account = frappe.db.get_value("Bank Account", bank_account, "account")
        if not account:
            return 0.0
        cutoff = add_days(getdate(from_date), -1)
        return _gl_balance(account, cutoff)
    except Exception:
        return 0.0


@frappe.whitelist()
def get_balance_summary(bank_account, from_date, to_date, company=None):
    """Return ERP GL closing balance at to_date — matches the
    'Closing (Opening + Total)' row of the General Ledger report."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    try:
        gl_account = frappe.db.get_value("Bank Account", bank_account, "account")
        if not gl_account:
            return {"erp_closing": 0.0}
        erp_closing = _gl_balance(gl_account, to_date)
        return {"erp_closing": erp_closing}
    except Exception:
        return {"erp_closing": 0.0}


@frappe.whitelist()
def get_erp_vouchers_for_match(bank_transaction, preselected_entry=None):
    """Return all submitted ERP vouchers in a ±60-day window around this transaction.
    If preselected_entry is provided (the AI-suggested match), it is always included even if
    it falls outside the date window — so the user can always see and confirm the AI pick.
    No amount filtering — the user picks manually; client-side 'exact amount' filter handles narrowing."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    txn = frappe.db.get_value(
        "Bank Transaction", bank_transaction,
        ["deposit", "withdrawal", "date", "bank_account"],
        as_dict=True,
    )
    if not txn:
        return []

    txn_date = txn.date
    company  = frappe.db.get_value("Bank Account", txn.bank_account, "company") or ""
    from frappe.utils import add_days
    date_from = add_days(txn_date, -60)
    date_to   = add_days(txn_date,  60)
    txn_date_obj = getdate(txn_date)

    vouchers = []

    pe_filters = [
        ["posting_date", "between", [date_from, date_to]],
        ["docstatus", "=", 1],
    ]
    if company:
        pe_filters.append(["company", "=", company])

    # Payment Entries — submitted, within date window, same company
    pes = frappe.db.get_all(
        "Payment Entry",
        filters=pe_filters,
        fields=["name", "posting_date", "payment_type", "party", "party_name",
                "paid_amount", "received_amount", "reference_no"],
        order_by="posting_date desc",
    )
    for pe in pes:
        pe_amount = float(pe.paid_amount or 0) or float(pe.received_amount or 0)
        vouchers.append({
            "name":         pe.name,
            "type":         "Payment Entry",
            "date":         str(pe.posting_date),
            "party":        pe.party_name or pe.party or "",
            "amount":       pe_amount,
            "reference":    pe.reference_no or "",
            "payment_type": pe.payment_type or "",
        })

    je_filters = [
        ["posting_date", "between", [date_from, date_to]],
        ["docstatus", "=", 1],
    ]
    if company:
        je_filters.append(["company", "=", company])

    # Journal Entries — submitted, within date window, same company
    jes = frappe.db.get_all(
        "Journal Entry",
        filters=je_filters,
        fields=["name", "posting_date", "total_debit", "total_credit", "cheque_no", "remark"],
        order_by="posting_date desc",
    )
    for je in jes:
        je_amount = float(je.total_debit or 0) or float(je.total_credit or 0)
        vouchers.append({
            "name":         je.name,
            "type":         "Journal Entry",
            "date":         str(je.posting_date),
            "party":        (je.remark or "")[:60],
            "amount":       je_amount,
            "reference":    je.cheque_no or "",
            "payment_type": "",
        })

    pi_filters = [
        ["posting_date", "between", [date_from, date_to]],
        ["docstatus", "=", 1],
    ]
    if company:
        pi_filters.append(["company", "=", company])

    # Purchase Invoices — submitted, within date window, same company
    pis = frappe.db.get_all(
        "Purchase Invoice",
        filters=pi_filters,
        fields=["name", "posting_date", "supplier", "supplier_name", "grand_total", "bill_no"],
        order_by="posting_date desc",
    )
    for pi in pis:
        vouchers.append({
            "name":         pi.name,
            "type":         "Purchase Invoice",
            "date":         str(pi.posting_date),
            "party":        pi.supplier_name or pi.supplier or "",
            "amount":       float(pi.grand_total or 0),
            "reference":    pi.bill_no or "",
            "payment_type": "",
        })

    si_filters = [
        ["posting_date", "between", [date_from, date_to]],
        ["docstatus", "=", 1],
    ]
    if company:
        si_filters.append(["company", "=", company])

    # Sales Invoices — submitted, within date window, same company
    sis = frappe.db.get_all(
        "Sales Invoice",
        filters=si_filters,
        fields=["name", "posting_date", "customer", "customer_name", "grand_total"],
        order_by="posting_date desc",
    )
    for si in sis:
        vouchers.append({
            "name":         si.name,
            "type":         "Sales Invoice",
            "date":         str(si.posting_date),
            "party":        si.customer_name or si.customer or "",
            "amount":       float(si.grand_total or 0),
            "reference":    "",
            "payment_type": "",
        })

    # Every OTHER doctype that actually posted against this bank's GL account —
    # Expense Claim, Loan Disbursement/Repayment, and anything else a given site
    # books through the bank. Discovered via GL Entry rather than by hardcoding
    # a query (and a bank-account field name) per doctype, so this covers
    # whatever a site actually uses without needing those apps installed.
    # Payment Entry and Journal Entry are already fetched above and excluded
    # here — note a "Contra Entry" is just a Journal Entry voucher_type, so it
    # is already included by that fetch.
    gl_bank_account = frappe.db.get_value("Bank Account", txn.bank_account, "account")
    if gl_bank_account:
        gl_rows = frappe.db.get_all(
            "GL Entry",
            filters={
                "account":      gl_bank_account,
                "posting_date": ["between", [date_from, date_to]],
                "is_cancelled": 0,
                "voucher_type": ["not in", ["Payment Entry", "Journal Entry"]],
            },
            fields=["voucher_type", "voucher_no", "posting_date", "debit", "credit", "party"],
        )
        # One voucher can post several lines against the bank account; net them
        # so each voucher appears once with its true bank-side movement.
        # Debit and credit must be netted, not added: a voucher carrying both a
        # debit and a credit line against the bank (a correction, or a transfer
        # booked in one voucher) would otherwise report the sum of the two legs
        # instead of the amount that actually moved, and so never line up with
        # the bank transaction's figure.
        other_vouchers = {}
        for gl in gl_rows:
            key = (gl.voucher_type, gl.voucher_no)
            agg = other_vouchers.setdefault(key, {"net": 0.0, "date": gl.posting_date, "party": gl.party or ""})
            agg["net"] += float(gl.debit or 0) - float(gl.credit or 0)
            if not agg["party"] and gl.party:
                agg["party"] = gl.party
        for (vtype, vno), agg in other_vouchers.items():
            amount = abs(agg["net"])
            # A voucher whose bank-side legs cancel out moved nothing through
            # the bank; there is no amount for a bank line to match against.
            if not amount:
                continue
            vouchers.append({
                "name":         vno,
                "type":         vtype,
                "date":         str(agg["date"] or ""),
                "party":        agg["party"],
                "amount":       amount,
                "reference":    "",
                "payment_type": "",
            })

    # If the AI suggested an invoice and a submitted PE already references it, surface that PE instead.
    # This covers the case where user created the PE (via "+ Create PE"), submitted it, and came back.
    if preselected_entry and (
        frappe.db.exists("Sales Invoice", preselected_entry) or
        frappe.db.exists("Purchase Invoice", preselected_entry)
    ):
        pe_refs = frappe.db.get_all(
            "Payment Entry Reference",
            filters={"reference_name": preselected_entry},
            fields=["parent"],
        )
        for per in pe_refs:
            if frappe.db.get_value("Payment Entry", per.parent, "docstatus") == 1:
                preselected_entry = per.parent
                break

    # If an AI-suggested entry was specified but isn't in the ±60-day window, fetch and inject it.
    if preselected_entry:
        existing_names = {v["name"] for v in vouchers}
        if preselected_entry not in existing_names:
            _inject_preselected(preselected_entry, vouchers)

    # Sort all vouchers by date proximity to the bank transaction, but float
    # amount matches to the top first.
    #
    # The list is truncated to 100 below, and the modal's "exact amount" filter
    # runs client-side over only what is returned. Sorting by date alone, a
    # voucher for exactly this amount a few weeks out could be cut by 100
    # same-week vouchers of unrelated amounts — and the user filtering by exact
    # amount would then be told there is no match at all. Widening the doctypes
    # searched (Expense Claim and friends, via the GL sweep above) puts more
    # vouchers in competition for those 100 slots, so make the ones the user is
    # actually hunting for immune to the cut.
    txn_amount = abs(float(txn.deposit or 0) - float(txn.withdrawal or 0))

    def _sort_key(v):
        amount_gap = abs(float(v.get("amount") or 0) - txn_amount)
        # 0.01 absorbs currency rounding; anything closer counts as the amount.
        is_amount_match = 0 if (txn_amount and amount_gap <= 0.01) else 1
        return (is_amount_match, abs((getdate(v["date"]) - txn_date_obj).days))

    vouchers.sort(key=_sort_key)

    # Pin the preselected entry to position 0 regardless of where the date sort placed it.
    # This ensures the AI-suggested match is always the first row the user sees in the modal.
    if preselected_entry:
        idx = next((i for i, v in enumerate(vouchers) if v["name"] == preselected_entry), None)
        if idx is not None and idx != 0:
            vouchers.insert(0, vouchers.pop(idx))

    return vouchers[:100]


@frappe.whitelist()
def search_erp_vouchers(bank_transaction, query=None, amount_tolerance=None,
                        date_from=None, date_to=None, limit=100):
    """Free-text search across ERP vouchers for the manual-match window.

    get_erp_vouchers_for_match deliberately returns a browsable default list:
    a +/-60-day window around the bank date, capped at 100 rows. That is the
    right default, but it means when the AI's suggestion is wrong the correct
    voucher may never reach the browser at all - outside the window, or cut by
    the cap - and no amount of client-side filtering can find what was never
    sent. This endpoint is the escape hatch: it searches server-side with no
    date window (unless one is passed explicitly), so a reviewer can reach any
    voucher by number, party, reference or amount.

    Returns the same row shape as get_erp_vouchers_for_match, plus two extra
    flags the UI uses to stop users picking a voucher that cannot work:
      can_reconcile - False for invoices and GL-derived vouchers, which
                      approve_match cannot clear directly.
      reconciled    - True when the voucher already has a clearance_date, i.e.
                      it is already matched to some bank transaction.
    """
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    txn = frappe.db.get_value(
        "Bank Transaction", bank_transaction,
        ["deposit", "withdrawal", "date", "bank_account"],
        as_dict=True,
    )
    if not txn:
        return []

    q = (query or "").strip()
    if not q:
        # Empty search is not "match everything" - the caller shows the default
        # list instead, so returning nothing here keeps the two paths distinct.
        return []

    try:
        limit = max(1, min(int(limit or 100), 200))
    except (TypeError, ValueError):
        limit = 100

    company = frappe.db.get_value("Bank Account", txn.bank_account, "company") or ""

    # Escape LIKE wildcards so a voucher number containing _ or % is searched
    # literally rather than as a pattern.
    like = "%" + q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_") + "%"

    # A numeric query also searches amounts, within an optional tolerance -
    # this is how a reviewer finds a near-miss (bank charge, FX difference,
    # partial payment) that an exact-amount filter would hide.
    try:
        amt = float(q.replace(",", ""))
    except (TypeError, ValueError):
        amt = None
    try:
        tol = abs(float(amount_tolerance or 0))
    except (TypeError, ValueError):
        tol = 0.0
    # Deliberately >= / <= rather than a "between" filter: Frappe's between
    # operator routes through its date handling and silently mangles a numeric
    # range into a single bogus value ("... paid_amount between 0.0"), which
    # MariaDB then rejects as a syntax error.
    amt_lo = (amt - tol) if amt is not None else None
    amt_hi = (amt + tol) if amt is not None else None

    def _base_filters():
        f = [["docstatus", "=", 1]]
        if company:
            f.append(["company", "=", company])
        # Only constrain dates when the caller explicitly asks; the whole point
        # of this endpoint is to escape the default window.
        if date_from and date_to:
            f.append(["posting_date", "between", [date_from, date_to]])
        elif date_from:
            f.append(["posting_date", ">=", date_from])
        elif date_to:
            f.append(["posting_date", "<=", date_to])
        return f

    def _search(doctype, fields, text_fields, amount_fields, builder):
        """Text search plus (optionally) an amount search, merged and deduped.

        The amount search has to be its own query rather than another
        or_filters clause: an or_filters entry is a single condition, and
        an amount window needs two (>= and <=) ANDed together.
        """
        found = {}
        for row in frappe.db.get_all(
            doctype,
            filters=_base_filters(),
            or_filters=[[f, "like", like] for f in text_fields],
            fields=fields, order_by="posting_date desc", limit_page_length=limit,
        ):
            found[row["name"]] = row
        if amt_lo is not None:
            for af in amount_fields:
                for row in frappe.db.get_all(
                    doctype,
                    filters=_base_filters() + [[af, ">=", amt_lo], [af, "<=", amt_hi]],
                    fields=fields, order_by="posting_date desc", limit_page_length=limit,
                ):
                    found.setdefault(row["name"], row)
        return [builder(r) for r in found.values()]

    results = []

    # ---- Payment Entries (directly reconcilable) ----
    results += _search(
        "Payment Entry",
        ["name", "posting_date", "payment_type", "party", "party_name",
         "paid_amount", "received_amount", "reference_no", "clearance_date"],
        ["name", "party", "party_name", "reference_no"],
        ["paid_amount", "received_amount"],
        lambda pe: {
            "name":          pe.name,
            "type":          "Payment Entry",
            "date":          str(pe.posting_date or ""),
            "party":         pe.party_name or pe.party or "",
            "amount":        float(pe.paid_amount or 0) or float(pe.received_amount or 0),
            "reference":     pe.reference_no or "",
            "payment_type":  pe.payment_type or "",
            "can_reconcile": True,
            "reconciled":    bool(pe.clearance_date),
        },
    )

    # ---- Journal Entries (directly reconcilable) ----
    results += _search(
        "Journal Entry",
        ["name", "posting_date", "total_debit", "total_credit",
         "cheque_no", "remark", "clearance_date"],
        ["name", "cheque_no", "remark"],
        ["total_debit", "total_credit"],
        lambda je: {
            "name":          je.name,
            "type":          "Journal Entry",
            "date":          str(je.posting_date or ""),
            "party":         (je.remark or "")[:60],
            "amount":        float(je.total_debit or 0) or float(je.total_credit or 0),
            "reference":     je.cheque_no or "",
            "payment_type":  "",
            "can_reconcile": True,
            "reconciled":    bool(je.clearance_date),
        },
    )

    # ---- Invoices (searchable, but NOT directly reconcilable) ----
    # Surfaced because reviewers search by invoice number constantly; flagged
    # can_reconcile False so the UI can show why it cannot be picked, rather
    # than letting approve_match throw after the fact.
    results += _search(
        "Purchase Invoice",
        ["name", "posting_date", "supplier", "supplier_name", "grand_total", "bill_no"],
        ["name", "supplier", "supplier_name", "bill_no"],
        ["grand_total"],
        lambda pi: {
            "name":          pi.name,
            "type":          "Purchase Invoice",
            "date":          str(pi.posting_date or ""),
            "party":         pi.supplier_name or pi.supplier or "",
            "amount":        float(pi.grand_total or 0),
            "reference":     pi.bill_no or "",
            "payment_type":  "",
            "can_reconcile": False,
            "reconciled":    False,
        },
    )

    results += _search(
        "Sales Invoice",
        ["name", "posting_date", "customer", "customer_name", "grand_total"],
        ["name", "customer", "customer_name"],
        ["grand_total"],
        lambda si: {
            "name":          si.name,
            "type":          "Sales Invoice",
            "date":          str(si.posting_date or ""),
            "party":         si.customer_name or si.customer or "",
            "amount":        float(si.grand_total or 0),
            "reference":     "",
            "payment_type":  "",
            "can_reconcile": False,
            "reconciled":    False,
        },
    )

    # ---- Other doctypes that posted to this bank's GL account ----
    # Kept consistent with the default list, which shows these too. Matched on
    # voucher number only (GL Entry carries no party_name or reference).
    gl_bank_account = frappe.db.get_value("Bank Account", txn.bank_account, "account")
    if gl_bank_account:
        gl_filters = [
            ["account", "=", gl_bank_account],
            ["is_cancelled", "=", 0],
            ["voucher_type", "not in", ["Payment Entry", "Journal Entry"]],
            ["voucher_no", "like", like],
        ]
        if date_from and date_to:
            gl_filters.append(["posting_date", "between", [date_from, date_to]])
        gl_agg = {}
        for gl in frappe.db.get_all(
            "GL Entry", filters=gl_filters,
            fields=["voucher_type", "voucher_no", "posting_date", "debit", "credit", "party"],
            limit_page_length=limit * 4,
        ):
            key = (gl.voucher_type, gl.voucher_no)
            a = gl_agg.setdefault(key, {"net": 0.0, "date": gl.posting_date, "party": gl.party or ""})
            a["net"] += float(gl.debit or 0) - float(gl.credit or 0)
            if not a["party"] and gl.party:
                a["party"] = gl.party
        for (vtype, vno), a in gl_agg.items():
            amount = abs(a["net"])
            if not amount:
                continue
            results.append({
                "name":          vno,
                "type":          vtype,
                "date":          str(a["date"] or ""),
                "party":         a["party"],
                "amount":        amount,
                "reference":     "",
                "payment_type":  "",
                "can_reconcile": False,
                "reconciled":    False,
            })

    # Usable rows first (reconcilable and not already cleared), then closest to
    # the bank transaction's own amount, then newest - so the row the reviewer
    # most likely wants is at the top and never lost to the result cap.
    bank_amount = abs(float(txn.deposit or 0) - float(txn.withdrawal or 0))

    # Two passes, relying on Python's stable sort: newest first, then rank on
    # top of that. Clearer than expressing a descending date inside a single
    # mixed sort key.
    results.sort(key=lambda v: v.get("date") or "", reverse=True)
    results.sort(key=lambda v: (
        0 if (v["can_reconcile"] and not v["reconciled"]) else 1,
        abs(float(v.get("amount") or 0) - bank_amount) if bank_amount else 0,
    ))
    return results[:limit]


def _inject_preselected(entry_name, vouchers):
    """Fetch a specific ERP entry by name and prepend it to the voucher list.
    Handles PE, JE, Sales Invoice, and Purchase Invoice — covers all entry types
    the AI can suggest so the preselected match is always visible in the modal."""
    # Try Payment Entry
    pe = frappe.db.get_value(
        "Payment Entry", entry_name,
        ["name", "posting_date", "payment_type", "party", "party_name", "paid_amount", "received_amount", "reference_no"],
        as_dict=True,
    )
    if pe:
        vouchers.insert(0, {
            "name":         pe.name,
            "type":         "Payment Entry",
            "date":         str(pe.posting_date or ""),
            "party":        pe.party_name or pe.party or "",
            "amount":       float(pe.paid_amount or 0) or float(pe.received_amount or 0),
            "reference":    pe.reference_no or "",
            "payment_type": pe.payment_type or "",
        })
        return
    # Try Journal Entry
    je = frappe.db.get_value(
        "Journal Entry", entry_name,
        ["name", "posting_date", "voucher_type", "total_debit", "total_credit", "cheque_no", "remark"],
        as_dict=True,
    )
    if je:
        vouchers.insert(0, {
            "name":         je.name,
            "type":         "Journal Entry",
            "date":         str(je.posting_date or ""),
            "party":        (je.remark or "")[:60],
            "amount":       float(je.total_debit or 0) or float(je.total_credit or 0),
            "reference":    je.cheque_no or "",
            "payment_type": je.voucher_type or "",
        })
        return
    # Try Sales Invoice
    si = frappe.db.get_value(
        "Sales Invoice", entry_name,
        ["name", "posting_date", "customer", "customer_name", "grand_total"],
        as_dict=True,
    )
    if si:
        vouchers.insert(0, {
            "name":         si.name,
            "type":         "Sales Invoice",
            "date":         str(si.posting_date or ""),
            "party":        si.customer_name or si.customer or "",
            "amount":       float(si.grand_total or 0),
            "reference":    "",
            "payment_type": "",
        })
        return
    # Try Purchase Invoice
    pi = frappe.db.get_value(
        "Purchase Invoice", entry_name,
        ["name", "posting_date", "supplier", "supplier_name", "grand_total", "bill_no"],
        as_dict=True,
    )
    if pi:
        vouchers.insert(0, {
            "name":         pi.name,
            "type":         "Purchase Invoice",
            "date":         str(pi.posting_date or ""),
            "party":        pi.supplier_name or pi.supplier or "",
            "amount":       float(pi.grand_total or 0),
            "reference":    pi.bill_no or "",
            "payment_type": "",
        })


def _consolidate_via_existing_match(txn_names, company=None, bank_account=None, label="Consolidated"):
    txns = frappe.db.get_all(
        "Bank Transaction",
        filters={"name": ["in", txn_names], "docstatus": 1},
        fields=["name", "date", "deposit", "withdrawal", "description",
                "reference_number", "party_type", "party", "bank_account"],
    )
    if not txns:
        frappe.throw(_("No valid transactions found."))

    # A consolidated group must be all-deposit or all-withdrawal — mixing the
    # two sides would silently net them against each other into one merged
    # amount (e.g. a deposit and a withdrawal cancelling out), which is never
    # a real single bank event and would misrepresent both transactions.
    txn_types = set()
    for t in txns:
        if float(t.deposit or 0) > 0:
            txn_types.add("deposit")
        if float(t.withdrawal or 0) > 0:
            txn_types.add("withdrawal")
    if len(txn_types) > 1:
        frappe.throw(_(
            "Cannot consolidate Deposit and Withdrawal transactions together. "
            "Select transactions that are all Deposits or all Withdrawals."
        ))

    total_deposit    = sum(float(t.deposit    or 0) for t in txns)
    total_withdrawal = sum(float(t.withdrawal or 0) for t in txns)
    used_account     = bank_account or txns[0].bank_account
    latest_date      = max(t.date for t in txns)
    net              = total_deposit - total_withdrawal

    company_name = company or frappe.db.get_value("Bank Account", used_account, "company") or \
                   frappe.defaults.get_defaults().get("company")


    parties = {(t.get("party") or "").strip() for t in txns if t.get("party")}
    common_party = next(iter(parties)) if len(parties) == 1 else ""

    virtual_txn = {
        "deposit":          net if net >= 0 else 0,
        "withdrawal":       abs(net) if net < 0 else 0,
        "date":             latest_date,
        "description":      label + ": " + ", ".join(t.name for t in txns),
        "reference_number": "",
        "party":            common_party,
        "party_type":       (txns[0].get("party_type") or "") if common_party else "",
    }

    from .signal_calculator import SignalCalculator
    from .pattern_store import PatternStore

    engine = BankMatchingEngine(used_account, latest_date, latest_date, company_name)
    candidates = [
        c for c in engine._get_candidates()
        if c.get("entry_type") in ("Payment Entry", "Journal Entry")
    ]

    settings = _load_sbr_settings()
    signal_calc = SignalCalculator(
        PatternStore(),
        amount_tolerance_pct=settings.get("amount_tolerance_pct"),
        date_window=settings.get("date_window_days"),
    )
    scored = signal_calc.score_all(virtual_txn, candidates)

    best = scored[0] if scored else None

    result = {
        "count":            len(txns),
        "total_deposit":    total_deposit,
        "total_withdrawal": total_withdrawal,
        "net_amount":       net,
        "matched":          best is not None,
    }

    if best is not None:
        queue = "Review"
        confidence = best["confidence"]
        matched_entries_json = frappe.as_json([best["name"]])
        reasoning = "{0} {1} transactions (net {2:,.2f}) by {3} — matched against {4} {5}. {6}".format(
            label, len(txns), net, frappe.session.user, best["entry_type"], best["name"], best.get("reasoning") or ""
        ).strip()
        result.update({
            "entry_type":    best["entry_type"],
            "matched_entry": best["name"],
            "confidence":    confidence,
        })
    else:
        # No existing ERP entry fits the combined amount — still group the
        # transactions; the user creates a voucher manually from any of them
        # if needed.
        queue = "Unmatched"
        confidence = 0
        matched_entries_json = None
        reasoning = "{0} {1} transactions (net {2:,.2f}) by {3} — no existing ERP entry matched.".format(
            label, len(txns), net, frappe.session.user
        )

    group_id = frappe.generate_hash(length=12)

    for txn in txns:
        frappe.db.set_value("Bank Transaction", txn.name, {
            "recon_queue":           queue,
            "recon_match_type":      "Consolidated",
            "recon_confidence":      confidence,
            "recon_matched_entries": matched_entries_json,
            "recon_ai_reasoning":    reasoning,
            "recon_run_id":          group_id,
            "recon_user_action":     "Consolidated",
        })
    frappe.db.commit()

    return result


@frappe.whitelist()
def consolidate_transactions(transaction_names, company=None):
    frappe.only_for(["Accounts Manager", "System Manager"])
    if isinstance(transaction_names, str):
        transaction_names = json.loads(transaction_names)
    if len(transaction_names) < 2:
        frappe.throw(_("Select at least 2 transactions to consolidate."))
    return _consolidate_via_existing_match(transaction_names, company=company, label="Consolidated")


@frappe.whitelist()
def get_consolidatable_transactions(bank_account, from_date, to_date):
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
            "status": ["!=", "Reconciled"],
            "unallocated_amount": [">", 0],
        },
        fields=["name", "date", "deposit", "withdrawal", "description", "reference_number",
                "unallocated_amount", "recon_matched_entries"],
        order_by="date asc",
    )
    txns = [t for t in txns if not t.get("recon_matched_entries")]
    for t in txns:
        t.pop("recon_matched_entries", None)
    return txns


@frappe.whitelist()
def consolidate_bank_charges(bank_account, amount, consolidation_type, charges_account=None, company=None):
    frappe.only_for(["Accounts Manager", "System Manager"])

    amount = float(amount or 0)
    if amount <= 0:
        frappe.throw(_("Amount must be greater than 0."))
    if consolidation_type not in ("Debit", "Credit"):
        frappe.throw(_("Consolidation Type must be Debit or Credit."))

    gl_account = frappe.db.get_value("Bank Account", bank_account, "account")
    if not gl_account:
        frappe.throw(_("The selected Bank Account has no linked GL Account."))

    company_name = company or frappe.db.get_value("Account", gl_account, "company") or \
                   frappe.defaults.get_defaults().get("company")

    # Auto-discover a bank charges / interest account if not supplied
    if not charges_account:
        root_type = "Expense" if consolidation_type == "Debit" else "Income"
        for keyword in ["%Bank Charge%", "%Bank Fee%", "%Service Charge%", "%Interest%"]:
            charges_account = frappe.db.get_value(
                "Account",
                {
                    "company": company_name,
                    "account_name": ["like", keyword],
                    "root_type": root_type,
                    "is_group": 0,
                },
                "name",
            ) or ""
            if charges_account:
                break

    if not charges_account:
        frappe.throw(_(
            "Cannot find a suitable account for {0}. "
            "Please create a 'Bank Charges' (Expense) or 'Interest Income' (Income) account "
            "in the Chart of Accounts for {1}."
        ).format(consolidation_type, company_name))

    narration = "Consolidated bank charges ({0})".format(consolidation_type)
    je = frappe.new_doc("Journal Entry")
    je.voucher_type  = "Bank Entry"
    je.posting_date  = nowdate()
    je.company       = company_name
    je.remark        = narration
    je.cheque_no     = "BANK-CHARGES"
    je.cheque_date   = nowdate()

    if consolidation_type == "Debit":
        # Money OUT – bank charged us: DR charges account, CR bank
        je.append("accounts", {"account": charges_account, "debit_in_account_currency":  amount, "credit_in_account_currency": 0, "user_remark": narration})
        je.append("accounts", {"account": gl_account,      "debit_in_account_currency":  0,      "credit_in_account_currency": amount, "user_remark": narration})
    else:
        # Money IN – interest / refund: DR bank, CR income account
        je.append("accounts", {"account": gl_account,      "debit_in_account_currency":  amount, "credit_in_account_currency": 0, "user_remark": narration})
        je.append("accounts", {"account": charges_account, "debit_in_account_currency":  0,      "credit_in_account_currency": amount, "user_remark": narration})

    frappe.db.commit()
    try:
        je.insert(ignore_permissions=True)
    except Exception as _exc:
        if "1020" in str(_exc) or "Record has changed" in str(_exc):
            import time as _time
            _time.sleep(0.1)
            je.insert(ignore_permissions=True)
        else:
            raise

    frappe.db.commit()
    return {"journal_entry": je.name, "amount": amount, "consolidation_type": consolidation_type}


@frappe.whitelist()
def get_bank_charge_transactions(bank_account, from_date, to_date):
    """Return unreconciled BTs identified as bank charges (keyword match OR debit ≤ threshold)."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    settings = _load_sbr_settings()
    threshold = flt(settings.get("bank_charge_amount_threshold", 2000))

    all_txns = frappe.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [getdate(from_date), getdate(to_date)]],
            "docstatus": 1,
            "status": ["!=", "Reconciled"],
        },
        fields=["name", "date", "deposit", "withdrawal", "description", "reference_number"],
        order_by="date asc",
    )

    result = []
    for txn in all_txns:
        description_upper = (txn.get("description") or "").upper()
        withdrawal = flt(txn.get("withdrawal"))

        matched_charge_type = None
        for rule in _SBR_BANK_CHARGE_RULES:
            for kw in rule["keywords"]:
                if kw.upper() in description_upper:
                    matched_charge_type = rule["charge_type"]
                    break
            if matched_charge_type:
                break

        if matched_charge_type:
            result.append({
                "name": txn["name"],
                "date": str(txn.get("date") or ""),
                "description": txn.get("description") or "",
                "deposit": flt(txn.get("deposit")),
                "withdrawal": withdrawal,
                "reference_number": txn.get("reference_number") or "",
                "charge_type": matched_charge_type,
                "matched_by": "Keyword",
            })
        elif withdrawal > 0 and withdrawal <= threshold:
            result.append({
                "name": txn["name"],
                "date": str(txn.get("date") or ""),
                "description": txn.get("description") or "",
                "deposit": flt(txn.get("deposit")),
                "withdrawal": withdrawal,
                "reference_number": txn.get("reference_number") or "",
                "charge_type": "Other Bank Charge",
                "matched_by": "Amount ≤ ₦{:,.0f}".format(threshold),
            })

    return result


@frappe.whitelist()
def consolidate_selected_bank_charges(transaction_names, bank_account=None, company=None):
    """Combine selected bank-charge transactions and search for a matching
    existing ERP entry. See _consolidate_via_existing_match for the full
    rationale — this creates no new document."""
    frappe.only_for(["Accounts Manager", "System Manager"])

    if isinstance(transaction_names, str):
        transaction_names = json.loads(transaction_names)

    if len(transaction_names) < 2:
        frappe.throw(_("Select at least 2 transactions to consolidate."))

    return _consolidate_via_existing_match(
        transaction_names, company=company, bank_account=bank_account,
        label="Bank charges consolidated",
    )


@frappe.whitelist()
def get_erp_vouchers(bank_account, from_date, to_date):
    """Fetch submitted Payment Entries and Journal Entries for the ERP Vouchers tab."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    ba_doc = frappe.db.get_value("Bank Account", bank_account, ["account", "company"], as_dict=True) or {}
    gl_account = ba_doc.get("account")
    company    = ba_doc.get("company")
    vouchers = []

    # Payment Entries scoped to the bank GL account (paid_from OR paid_to = gl_account)
    if gl_account:
        company_clause = "AND pe.company = %(company)s" if company else ""
        pe_rows = frappe.db.sql("""
            SELECT pe.name, pe.posting_date, pe.payment_type, pe.party_type, pe.party,
                   pe.paid_amount, pe.received_amount, pe.reference_no, pe.clearance_date
            FROM `tabPayment Entry` pe
            WHERE (pe.paid_from = %(gl)s OR pe.paid_to = %(gl)s)
              AND pe.posting_date BETWEEN %(fd)s AND %(td)s
              AND pe.docstatus = 1
              {cc}
            ORDER BY pe.posting_date DESC
        """.format(cc=company_clause), {
            "gl": gl_account, "fd": from_date, "td": to_date, "company": company or ""
        }, as_dict=True)
    else:
        pe_rows = []

    for pe in pe_rows:
        amount = float(pe.paid_amount or 0) or float(pe.received_amount or 0)
        vouchers.append({
            "name":         pe.name,
            "type":         "Payment Entry",
            "type_short":   "PE",
            "date":         str(pe.posting_date),
            "party":        pe.party or "",
            "amount":       amount,
            "reference":    pe.reference_no or "",
            "payment_type": pe.payment_type or "",
            "status":       "Cleared" if pe.clearance_date else "Unreconciled",
        })

    # Journal Entries linked to the bank GL account
    if gl_account:
        je_names_raw = frappe.db.get_all(
            "Journal Entry Account",
            filters={"account": gl_account, "docstatus": 1},
            fields=["parent"],
            distinct=True,
        )
        je_name_list = [j.parent for j in je_names_raw]
        if je_name_list:
            jes = frappe.db.get_all(
                "Journal Entry",
                filters=[
                    ["name", "in", je_name_list],
                    ["posting_date", "between", [from_date, to_date]],
                    ["docstatus", "=", 1],
                ],
                fields=["name", "posting_date", "voucher_type", "total_debit",
                        "cheque_no", "remark", "clearance_date"],
                order_by="posting_date desc",
            )
            for je in jes:
                vouchers.append({
                    "name":         je.name,
                    "type":         "Journal Entry",
                    "type_short":   "JE",
                    "date":         str(je.posting_date),
                    "party":        je.remark or "",
                    "amount":       float(je.total_debit or 0),
                    "reference":    je.cheque_no or "",
                    "payment_type": je.voucher_type or "",
                    "status":       "Cleared" if je.clearance_date else "Unreconciled",
                })

    vouchers.sort(key=lambda v: v["date"], reverse=True)
    return {"vouchers": vouchers, "total": len(vouchers)}


@frappe.whitelist()
def reset_bank_transactions(bank_account, from_date, to_date):
    """Cancel and delete all Bank Transactions for an account in a date range.
    Used by the demo data reset flow to clear accumulated duplicates before reimport."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    names = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [getdate(from_date), getdate(to_date)]],
        },
        pluck="name",
    )
    if not names:
        return {"deleted": 0}

    in_clause = ", ".join(["%s"] * len(names))
    frappe.db.sql(
        "UPDATE `tabBank Transaction` SET docstatus=2 WHERE name IN (" + in_clause + ")",
        names,
    )
    frappe.db.sql(
        "DELETE FROM `tabBank Transaction` WHERE name IN (" + in_clause + ")",
        names,
    )
    frappe.db.commit()
    return {"deleted": len(names)}


@frappe.whitelist()
def reset_ai_suggestions(bank_account, from_date, to_date):
    """Clear AI scoring fields for all unreconciled transactions in the period.
    Leaves truly-Reconciled (status=Reconciled) rows untouched."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    frappe.db.sql(
        "UPDATE `tabBank Transaction` SET "
        "recon_queue=NULL, recon_confidence=0, recon_matched_entries=NULL, "
        "recon_match_type=NULL, recon_run_id=NULL, recon_ai_reasoning=NULL, "
        "recon_draft_payload=NULL, recon_signals_json=NULL, recon_wht_amount=0, "
        "recon_user_action=NULL "
        "WHERE bank_account=%s AND `date` BETWEEN %s AND %s "
        "AND status != 'Reconciled' AND docstatus=1",
        (bank_account, from_date, to_date),
    )
    frappe.db.commit()
    return {"reset": True}


def _parse_flexible_date(date_str):
    """Parse bank statement dates that may use YYYY-DD-MM or other non-standard formats.

    Some Nigerian bank exports produce YYYY-DD-MM (e.g. 2026-13-01 meaning Jan 13).
    We try standard ISO / dayfirst formats first; if the 'month' slot > 12 we swap
    day and month.  Raises ValueError (not frappe.throw) so callers can silently skip.
    """
    import re as _re
    from datetime import date as _date
    import dateutil.parser as _dup

    if not date_str:
        return None
    date_str = str(date_str).strip()

    # Standard ISO (YYYY-MM-DD) and unambiguous formats
    try:
        return _dup.parse(date_str, dayfirst=False).date()
    except Exception:
        pass

    # Day-first formats: DD/MM/YYYY, DD-MM-YYYY
    try:
        return _dup.parse(date_str, dayfirst=True).date()
    except Exception:
        pass

    # YYYY-DD-MM: "month" slot > 12 means day and month are swapped
    m = _re.match(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$', date_str)
    if m:
        y, part2, part3 = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if part2 > 12 and 1 <= part3 <= 12:
            try:
                return _date(y, part3, part2)
            except ValueError:
                pass

    raise ValueError('Cannot parse date: {0}'.format(date_str))


@frappe.whitelist()
def import_bank_statement(bank_account, rows, company=None):
    """Parse JSON rows (from CSV import) and create submitted Bank Transactions."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import json
    if isinstance(rows, str):
        rows = json.loads(rows)
    if not rows:
        frappe.throw(_("No data rows provided."))

    ba = frappe.db.get_value("Bank Account", bank_account, ["bank", "company"], as_dict=True)
    company_name = company or (ba.company if ba else None) or \
                   frappe.defaults.get_defaults().get("company")

    # Pre-fetch existing transactions in the CSV's date range to skip duplicates
    all_dates = []
    for row in rows:
        d = (row.get("date") or row.get("Date") or "").strip()
        if d:
            try:
                all_dates.append(_parse_flexible_date(d))
            except Exception:
                pass

    existing_set = set()
    if all_dates:
        existing_txns = frappe.db.get_all(
            "Bank Transaction",
            filters={
                "bank_account": bank_account,
                "date": ["between", [min(all_dates), max(all_dates)]],
                "docstatus": 1,
            },
            fields=["date", "deposit", "withdrawal", "description"],
        )
        existing_set = {
            (str(t.date), float(t.deposit or 0), float(t.withdrawal or 0), (t.description or "").strip())
            for t in existing_txns
        }

    skipped = 0
    pending = []  # (name, unallocated_amount) for bulk submit
    for row in rows:
        try:
            date_val    = (row.get("date") or row.get("Date") or "").strip() or nowdate()
            credit      = float(row.get("credit") or row.get("Credit") or
                                row.get("deposit") or 0)
            debit       = float(row.get("debit") or row.get("Debit") or
                                row.get("withdrawal") or 0)
            desc        = (row.get("description") or row.get("Description") or
                           row.get("narration") or row.get("Narration") or "").strip()
            ref         = (row.get("reference") or row.get("Reference") or "").strip()

            if not credit and not debit:
                skipped += 1
                continue

            row_key = (str(_parse_flexible_date(date_val)), float(credit), float(debit), desc)
            if row_key in existing_set:
                skipped += 1
                continue

            bt = frappe.new_doc("Bank Transaction")
            bt.bank_account     = bank_account
            bt.date             = _parse_flexible_date(date_val)
            bt.deposit          = credit
            bt.withdrawal       = debit
            bt.description      = desc
            bt.reference_number = ref
            bt.company          = company_name
            bt.flags.ignore_links = True
            bt.insert(ignore_permissions=True)
            pending.append((bt.name, credit or debit))
        except Exception:
            frappe.log_error(frappe.get_traceback(), "import_bank_statement row failed")
            skipped += 1

    # Bulk-submit all inserted rows in one SQL — avoids N individual submit lifecycles
    if pending:
        in_clause = ", ".join(["%s"] * len(pending))
        case_when = " ".join(["WHEN %s THEN %s"] * len(pending))
        case_params = []
        for name, amt in pending:
            case_params.extend([name, amt])
        frappe.db.sql(
            "UPDATE `tabBank Transaction` SET docstatus=1, status='Unreconciled',"
            " unallocated_amount = CASE name " + case_when + " ELSE 0 END"
            " WHERE name IN (" + in_clause + ")",
            case_params + [n for n, _ in pending]
        )

    created = [n for n, _ in pending]
    frappe.db.commit()
    return {"created": len(created), "skipped": skipped, "names": created}


@frappe.whitelist()
def import_and_analyze(bank_account, rows, company, from_date, to_date):
    """Import bank statement rows then immediately run the matching engine.
    Returns import result + full suggestions payload in a single round-trip."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    import_result = import_bank_statement(bank_account, rows, company)
    suggestions = get_suggestions(bank_account, from_date, to_date, company)
    return {
        "import": import_result,
        "queue_counts": suggestions["queue_counts"],
        "transactions": suggestions["transactions"],
        "suggestions": suggestions["suggestions"],
    }


@frappe.whitelist()
def parse_statement_file(file_b64, filename):
    """Parse an uploaded bank statement (XLSX, XLS, or MT940/STA) server-side.

    CSV is parsed entirely client-side; this endpoint handles binary formats.
    Returns {rows, headers, count} — the caller passes rows to import_and_analyze.
    """
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    import base64

    try:
        raw = base64.b64decode(file_b64)
    except Exception:
        frappe.throw(_("Invalid file encoding — could not base64-decode the upload."))

    name_lower = (filename or "").lower()

    if name_lower.endswith(".xlsx"):
        rows, headers = _parse_excel_xlsx(raw)
    elif name_lower.endswith(".xls"):
        rows, headers = _parse_excel_xls(raw)
    elif any(name_lower.endswith(ext) for ext in (".mt940", ".sta", ".940", ".mt9")):
        rows = _parse_mt940(raw.decode("utf-8", errors="replace"))
        headers = ["date", "description", "debit", "credit", "reference"]
    else:
        frappe.throw(_("Unsupported format. Upload CSV, XLSX, or MT940/STA."))

    return {"rows": rows, "headers": headers, "count": len(rows)}


def _parse_excel_xlsx(raw_bytes):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    return _excel_rows_to_dicts(all_rows)


def _parse_excel_xls(raw_bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=raw_bytes)
    ws = wb.sheet_by_index(0)
    all_rows = [
        tuple(ws.cell(r, c).value for c in range(ws.ncols))
        for r in range(ws.nrows)
    ]
    return _excel_rows_to_dicts(all_rows)


_HEADER_DATE_NAMES = {"date", "value date", "trans date", "tran date", "transaction date"}
_HEADER_AMOUNT_NAMES = {"debit", "credit", "withdrawal", "withdrawals", "deposit", "deposits", "amount"}


def _looks_like_header_row(cells):
    """True only for a row that actually names a date column and an amount
    column — used to find the real header row. Some banks prepend a
    grand-totals line (mostly blank cells plus a few summary figures) above
    the real header, which blindly trusting row 0 would misread as headers,
    corrupting every column mapping below it."""
    norm = {str(c or "").strip().lower() for c in cells}
    return bool(norm & _HEADER_DATE_NAMES) and bool(norm & _HEADER_AMOUNT_NAMES)


def _excel_rows_to_dicts(all_rows):
    if not all_rows:
        return [], []
    import datetime as _dt

    header_idx = 0
    for i, row in enumerate(all_rows[: min(len(all_rows), 10)]):
        if _looks_like_header_row(row):
            header_idx = i
            break

    headers = [str(h or "").strip() for h in all_rows[header_idx]]
    rows = []
    for raw_row in all_rows[header_idx + 1 :]:
        vals = []
        for v in raw_row:
            if isinstance(v, (_dt.datetime, _dt.date)):
                vals.append(v.strftime("%Y-%m-%d"))
            elif v is not None:
                vals.append(str(v))
            else:
                vals.append("")
        if all(not v for v in vals):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows, headers


def _parse_mt940(text):
    """Minimal MT940 (SWIFT) parser — covers the standard :61:/:86: block structure."""
    import re
    rows = []
    blocks = re.split(r'(?=:61:)', text)
    for block in blocks:
        m61 = re.search(
            r':61:\s*(\d{6})(?:\d{0,4})([DC])[A-Z]?([\d,]+)N(\S*)',
            block,
        )
        if not m61:
            continue
        val_date_str = m61.group(1)
        dc = m61.group(2)
        amount_str = m61.group(3).replace(",", ".")
        ref = m61.group(4)
        try:
            year = 2000 + int(val_date_str[:2])
            month = int(val_date_str[2:4])
            day = int(val_date_str[4:6])
            date_val = f"{year:04d}-{month:02d}-{day:02d}"
        except (ValueError, IndexError):
            continue
        try:
            amount = float(amount_str)
        except ValueError:
            continue
        m86 = re.search(r':86:(.*?)(?=:\d{2}[A-Z]?:|$)', block, re.DOTALL)
        narration = ""
        if m86:
            narration = re.sub(r'\s+', ' ', m86.group(1)).strip()
            narration = re.sub(r'/[A-Z]{2,6}/', ' ', narration).strip()
        rows.append({
            "date": date_val,
            "description": narration or ref,
            "reference": ref,
            "debit": str(amount if dc == "D" else 0.0),
            "credit": str(amount if dc == "C" else 0.0),
        })
    return rows


@frappe.whitelist()
def get_reconciliation_report(bank_account, from_date, to_date, company=None):
    """Return reconciliation statistics and full transaction list for the report modal."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    # Resolve company from bank account if not provided
    if not company:
        ba = frappe.db.get_value("Bank Account", bank_account, ["company", "bank"], as_dict=True)
        company = (ba.company if ba else None) or frappe.defaults.get_defaults().get("company")
        bank_label = (ba.bank if ba else None) or bank_account
    else:
        ba = frappe.db.get_value("Bank Account", bank_account, ["bank"], as_dict=True)
        bank_label = (ba.bank if ba else None) or bank_account

    txns = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [getdate(from_date), getdate(to_date)]],
            "docstatus": 1,
        },
        fields=["name", "date", "deposit", "withdrawal", "description",
                "reference_number", "party", "status", "recon_queue",
                "recon_confidence", "recon_matched_entries", "recon_match_type",
                "recon_run_id", "unallocated_amount"],
        order_by="date asc",
    )

    import json as _json
    from collections import Counter
    # Counted per raw bank statement line, matching the main screen's tiles
    # (see _tally_queue_counts) — a consolidated group counts as its member
    # transactions, not as one.
    total        = len(txns)
    queue_counts = Counter(
        t.get("recon_queue") if t.get("recon_queue") in _KNOWN_RECON_QUEUES else "Unmatched"
        for t in txns
    )

    reconciled  = queue_counts.get("Reconciled", 0)
    auto        = queue_counts.get("Auto", 0)
    # Raw per-queue count — matches the main screen's own Review tile exactly
    # (previously combined with High-Val here, which made this number diverge
    # from the tile the user actually sees on the main screen).
    review      = queue_counts.get("Review", 0)
    unmatched   = queue_counts.get("Unmatched", 0)
    high_val    = queue_counts.get("High-Val", 0)
    dupes       = queue_counts.get("Duplicate", 0)
    aging       = queue_counts.get("Aging", 0)

    non_recon       = total - reconciled
    automation_rate = round((auto / non_recon * 100) if non_recon > 0 else 0, 1)

    # Manual vs auto reconciled breakdown
    auto_thresh = _load_sbr_settings().get("auto_threshold", 80)
    auto_reconciled   = sum(1 for t in txns
                            if t.get("recon_queue") == "Reconciled"
                            and float(t.get("recon_confidence") or 0) >= auto_thresh)
    manual_reconciled = reconciled - auto_reconciled
    manual_rate       = round((manual_reconciled / total * 100) if total > 0 else 0, 1)

    def _first_match(raw_json):
        if not raw_json:
            return ""
        try:
            entries = _json.loads(raw_json)
            return entries[0] if entries else ""
        except Exception:
            return ""

    # ERP entries with no clearance_date in the date range (PE + JE)
    erp_unmatched = []
    if company:
        pe_list = frappe.db.get_all(
            "Payment Entry",
            filters={
                "company": company,
                "docstatus": 1,
                "clearance_date": ["is", "not set"],
                "posting_date": ["between", [getdate(from_date), getdate(to_date)]],
                "payment_type": ["in", ["Receive", "Pay"]],
            },
            fields=["name", "posting_date", "payment_type", "party",
                    "paid_amount", "received_amount"],
            order_by="posting_date asc",
            limit=200,
        )
        for pe in pe_list:
            erp_unmatched.append({
                "name":       pe.name,
                "entry_type": "Payment Entry",
                "date":       str(pe.posting_date),
                "amount":     float(pe.received_amount or pe.paid_amount or 0),
                "party":      pe.party or "",
            })

        je_list = frappe.db.get_all(
            "Journal Entry",
            filters={
                "company": company,
                "docstatus": 1,
                "clearance_date": ["is", "not set"],
                "posting_date": ["between", [getdate(from_date), getdate(to_date)]],
            },
            fields=["name", "posting_date", "voucher_type", "total_debit", "total_credit"],
            order_by="posting_date asc",
            limit=200,
        )
        for je in je_list:
            erp_unmatched.append({
                "name":       je.name,
                "entry_type": je.voucher_type or "Journal Entry",
                "date":       str(je.posting_date),
                "amount":     float(je.total_debit or je.total_credit or 0),
                "party":      "",
            })

        erp_unmatched.sort(key=lambda x: x["date"])

    # For reconciled transactions missing recon_matched_entries, fall back to
    # Bank Transaction Payments child table to get the actual linked ERP entry.
    missing_match_names = [
        t.name for t in txns
        if t.get("recon_queue") == "Reconciled" and not t.get("recon_matched_entries")
    ]
    btp_lookup = {}
    if missing_match_names:
        btp_rows = frappe.db.get_all(
            "Bank Transaction Payments",
            filters={"parent": ["in", missing_match_names]},
            fields=["parent", "payment_entry"],
        )
        for row in btp_rows:
            if row.parent not in btp_lookup:
                btp_lookup[row.parent] = []
            btp_lookup[row.parent].append(row.payment_entry)

    def _resolve_match(t):
        match = _first_match(t.get("recon_matched_entries"))
        if not match and t.name in btp_lookup:
            entries = btp_lookup[t.name]
            match = entries[0] if entries else ""
        return match

    return {
        "period": {
            "from_date":    str(from_date),
            "to_date":      str(to_date),
            "bank_account": bank_account,
            "company":      company or "",
            "bank_label":   bank_label,
        },
        "summary": {
            "total":             total,
            "reconciled":        reconciled,
            "auto":              auto,
            "review":            review,
            "unmatched":         unmatched,
            "high_val":          high_val,
            "dupes":             dupes,
            "aging":             aging,
            "automation_rate":   automation_rate,
            "auto_reconciled":   auto_reconciled,
            "manual_reconciled": manual_reconciled,
            "manual_rate":       manual_rate,
        },
        "erp_unmatched": erp_unmatched,
        "transactions": [
            {
                "name":             t.name,
                "date":             str(t.date),
                "deposit":          float(t.deposit or 0),
                "withdrawal":       float(t.withdrawal or 0),
                "description":      (t.description or "")[:60],
                "reference_number": t.reference_number or "",
                "party":            t.party or "",
                "queue":            t.get("recon_queue") or "Unmatched",
                "confidence":       float(t.get("recon_confidence") or 0),
                "suggested_match":  _resolve_match(t),
            }
            for t in txns
        ],
    }


@frappe.whitelist()
def get_aging_erp_entries(bank_account, from_date=None, to_date=None, company=None):
    """Return submitted ERP vouchers (PE/JE) with no clearance_date that are older
    than the configured aging_days threshold — these are ERP entries waiting for a
    matching bank transaction (the S-19 / Aging-ERP scenario)."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    from frappe.utils import add_days, date_diff

    settings = _load_sbr_settings()
    aging_days = int(settings.get("aging_days", 10))
    cutoff = add_days(nowdate(), -aging_days)

    if not company:
        ba = frappe.db.get_value("Bank Account", bank_account, "company")
        company = ba or frappe.defaults.get_defaults().get("company")

    if not from_date:
        from_date = add_days(nowdate(), -365)
    if not to_date:
        to_date = nowdate()

    # Only look at entries that posted before the cutoff date AND within the requested range
    cutoff_d   = getdate(cutoff)
    from_date_d = getdate(from_date)
    to_date_d   = getdate(to_date)
    # effective window: from_date → min(cutoff, to_date)
    effective_to = cutoff_d if cutoff_d < to_date_d else to_date_d

    if effective_to < from_date_d:
        return {"entries": [], "count": 0, "aging_days": aging_days}

    entries = []

    pes = frappe.db.get_all(
        "Payment Entry",
        filters={
            "company": company,
            "docstatus": 1,
            "clearance_date": ["is", "not set"],
            "posting_date": ["between", [str(from_date_d), str(effective_to)]],
        },
        fields=["name", "posting_date", "payment_type", "party",
                "paid_amount", "received_amount", "reference_no"],
        limit=100,
    )
    for pe in pes:
        days_old = date_diff(nowdate(), pe.posting_date)
        entries.append({
            "name":       pe.name,
            "entry_type": "Payment Entry",
            "date":       str(pe.posting_date),
            "amount":     float(pe.received_amount or pe.paid_amount or 0),
            "party":      pe.party or "",
            "days_old":   days_old,
            "reference":  pe.reference_no or "",
        })

    jes = frappe.db.get_all(
        "Journal Entry",
        filters={
            "company": company,
            "docstatus": 1,
            "clearance_date": ["is", "not set"],
            "posting_date": ["between", [str(from_date_d), str(effective_to)]],
        },
        fields=["name", "posting_date", "voucher_type", "total_debit", "total_credit", "cheque_no"],
        limit=100,
    )
    for je in jes:
        days_old = date_diff(nowdate(), je.posting_date)
        entries.append({
            "name":       je.name,
            "entry_type": je.voucher_type or "Journal Entry",
            "date":       str(je.posting_date),
            "amount":     float(je.total_debit or je.total_credit or 0),
            "party":      "",
            "days_old":   days_old,
            "reference":  je.cheque_no or "",
        })

    entries.sort(key=lambda x: x["days_old"], reverse=True)
    return {"entries": entries, "count": len(entries), "aging_days": aging_days}


@frappe.whitelist()
def get_queue_summary(bank_account, from_date, to_date):
    """Fast count-only query for the summary tiles."""
    rows = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [from_date, to_date]],
            "docstatus": 1,
        },
        fields=["recon_queue", "recon_match_type", "recon_matched_entries", "recon_run_id"],
    )
    return _tally_queue_counts(rows)


# ── Settings helpers ──────────────────────────────────────────────────────────

_SBR_DEFAULTS = {
    "auto_threshold":               80.0,
    "review_threshold":             50.0,
    "high_val_threshold":           50_000_000.0,
    "aging_days":                   10,
    "amount_tolerance_pct":         1.0,
    "date_window_days":             5,
    "suspense_account":             "",
    "bank_charge_amount_threshold": 2000.0,
}

# Nigerian bank charge keyword rules — order matters (first match wins).
# Each keyword is checked as a case-insensitive substring of the BT description.
_SBR_BANK_CHARGE_RULES = [
    {
        "charge_type": "Account Maintenance",
        "keywords": [
            "ACCOUNT MAINTENANCE FEE", "ACCOUNT MAINTENANCE CHARGE",
            "ACCT MAINT FEE", "MONTHLY MAINTENANCE FEE", "MAINTENANCE CHARGE",
            "MONTHLY ACCT FEE", "MONTHLY SERVICE FEE",
        ],
    },
    {
        "charge_type": "COT",
        "keywords": ["COMMISSION ON TURNOVER", "COT CHARGE", "COT"],
    },
    {
        "charge_type": "SMS / Alert",
        "keywords": [
            "SMS ALERT FEE", "SMS ALERT CHARGE", "SMS CHARGE",
            "SMS NOTIFICATION FEE", "E-ALERT CHARGE", "ELECTRONIC NOTIFICATION FEE",
            "SMS ALERT",
        ],
    },
    {
        "charge_type": "Stamp Duty",
        "keywords": [
            "STAMP DUTY DEDUCTION", "STAMP DUTY LEVY", "STAMP DUTY CHG", "STAMP DUTY",
        ],
    },
    {
        "charge_type": "VAT",
        "keywords": [
            "VAT ON BANK CHARGES", "VAT ON CHARGES", "VAT ON CHARGE",
            "VAT ON CHG", "VALUE ADDED TAX",
        ],
    },
    {
        "charge_type": "Card Charges",
        "keywords": [
            "CARD MAINTENANCE FEE", "CARD MAINT FEE", "CARD ISSUANCE FEE",
            "CARD ANNUAL FEE", "DEBIT CARD FEE", "CARD FEE",
        ],
    },
    {
        "charge_type": "NIP / Interbank",
        "keywords": [
            "NIP TRANSACTION FEE", "INTERBANK TRANSFER FEE", "NIBSS FEE", "NIP CHARGE",
        ],
    },
    {
        "charge_type": "NEFT",
        "keywords": ["NEFT CHARGE", "NEFT FEE"],
    },
    {
        "charge_type": "RTGS",
        "keywords": ["RTGS CHARGE", "RTGS FEE"],
    },
    {
        "charge_type": "USSD",
        "keywords": ["USSD CHARGE", "USSD FEE", "*737# FEE", "*901# FEE"],
    },
    {
        "charge_type": "Internet Banking",
        "keywords": ["INTERNET BANKING FEE", "E-BANKING FEE", "DIGITAL BANKING FEE"],
    },
    {
        "charge_type": "WHT",
        "keywords": ["WITHHOLDING TAX", "WHT ON INT", "WHT DEDUCTION"],
    },
]


def _load_sbr_settings():
    out = {}
    for k, default in _SBR_DEFAULTS.items():
        raw = frappe.db.get_default("sbr_" + k)
        out[k] = type(default)(raw) if raw is not None else default
    return out


@frappe.whitelist()
def get_sbr_settings():
    """Return current AI match thresholds."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    return _load_sbr_settings()


@frappe.whitelist()
def save_sbr_settings(settings_json):
    """Persist threshold settings to frappe.defaults."""
    frappe.only_for(["Accounts Manager", "System Manager"])
    settings = json.loads(settings_json) if isinstance(settings_json, str) else settings_json
    for k in _SBR_DEFAULTS:
        if k in settings:
            frappe.db.set_default("sbr_" + k, settings[k])
    frappe.db.commit()
    return {"ok": True}


# ── Audit trail ───────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_audit_trail(bank_account, from_date, to_date):
    """Return all bank transactions that have had a user action recorded."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    rows = frappe.db.get_all(
        "Bank Transaction",
        filters={
            "bank_account": bank_account,
            "date": ["between", [getdate(from_date), getdate(to_date)]],
            "recon_user_action": ["not in", ["", "Pending", None]],
            "docstatus": ["!=", 2],
        },
        fields=[
            "name", "date", "deposit", "withdrawal", "description",
            "reference_number", "party",
            "recon_user_action", "recon_queue", "recon_confidence",
            "recon_matched_entries", "recon_match_type",
            "modified", "modified_by",
        ],
        order_by="modified desc",
    )
    return {"actions": rows, "total": len(rows)}


# ── Unreconcile Bank Transaction ──────────────────────────────────────────────

@frappe.whitelist()
def get_linked_payment_entries_for_bt(bank_transaction_name):
    """Return the vouchers currently linked to a Bank Transaction."""
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])
    return frappe.get_all(
        "Bank Transaction Payments",
        filters={"parent": bank_transaction_name},
        fields=["name", "payment_document", "payment_entry", "allocated_amount"],
    )


@frappe.whitelist()
def unreconcile_bank_transaction(bank_transaction_name, payment_entry_name):
    """
    Remove a specific voucher link from a Bank Transaction (bank-side unreconcile).

    Mirrors ERPNext v13 on_cancel logic for clearance_date handling.
    Does NOT touch GL Entries, Payment Ledger Entries, or Payment Entry References
    — the payment's accounting allocation to invoices remains intact.
    """
    frappe.only_for(["Accounts User", "Accounts Manager", "System Manager"])

    bt = frappe.get_doc("Bank Transaction", bank_transaction_name)

    if bt.docstatus != 1:
        frappe.throw(_("Only submitted Bank Transactions can be unreconciled."))

    matching = [pe for pe in bt.payment_entries if pe.payment_entry == payment_entry_name]
    if not matching:
        frappe.throw(_(
            "{0} is not linked to Bank Transaction {1}."
        ).format(payment_entry_name, bank_transaction_name))

    entry = matching[0]

    # Step 1: Clear clearance_date on the voucher — mirrors BankTransaction.clear_simple_entry
    _sbr_clear_clearance_date(entry)

    # Step 2: Remove the child row
    frappe.db.delete("Bank Transaction Payments", {"name": entry.name})

    # Step 3: Recalculate amounts from remaining rows
    bt.reload()
    allocated = sum(flt(pe.allocated_amount) for pe in bt.payment_entries)
    total = abs(flt(bt.withdrawal) - flt(bt.deposit))
    unallocated = total - allocated
    status = "Reconciled" if unallocated <= 0 else "Unreconciled"

    frappe.db.set_value("Bank Transaction", bank_transaction_name, {
        "allocated_amount": allocated,
        "unallocated_amount": unallocated,
        "status": status,
    })

    frappe.db.commit()

    return {
        "success": True,
        "bank_transaction": bank_transaction_name,
        "payment_entry": payment_entry_name,
        "allocated_amount": allocated,
        "unallocated_amount": unallocated,
        "status": status,
    }


def _sbr_clear_clearance_date(entry):
    """Clear clearance_date on a voucher being unlinked from a Bank Transaction."""
    doc_type = entry.payment_document
    doc_name = entry.payment_entry

    if doc_type in [
        "Payment Entry", "Journal Entry", "Purchase Invoice",
        "Expense Claim", "Loan Repayment", "Loan Disbursement",
    ]:
        if doc_type == "Payment Entry":
            payment_type = frappe.db.get_value("Payment Entry", doc_name, "payment_type")
            if payment_type == "Internal Transfer":
                # For Internal Transfer both legs must be reconciled for clearance_date to be set.
                # If this is the only linked BT, clearance_date was never meaningfully set — skip.
                linked_bts = frappe.get_all(
                    "Bank Transaction Payments",
                    filters={"payment_entry": doc_name},
                    fields=["parent"],
                )
                if len(linked_bts) < 2:
                    return
        frappe.db.set_value(doc_type, doc_name, "clearance_date", None)

    elif doc_type == "Sales Invoice":
        frappe.db.set_value(
            "Sales Invoice Payment",
            {"parenttype": "Sales Invoice", "parent": doc_name},
            "clearance_date",
            None,
        )
